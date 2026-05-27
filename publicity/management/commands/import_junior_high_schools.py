import os
from django.core.management.base import BaseCommand
from django.conf import settings
from openpyxl import load_workbook

from publicity.models import JuniorHighSchool, JuniorHighClass


def to_int(value):
    if value is None:
        return 0

    if value == "":
        return 0

    if value == "―":
        return 0

    if value == "-":
        return 0

    try:
        return int(value)
    except ValueError:
        return 0


class Command(BaseCommand):
    help = "中学校情報ExcelをDBに取り込みます"

    def handle(self, *args, **options):
        file_path = os.path.join(settings.BASE_DIR, "data", "junior_high_schools.xlsx")

        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"Excelファイルが見つかりません: {file_path}"))
            return

        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        JuniorHighClass.objects.all().delete()
        JuniorHighSchool.objects.all().delete()

        current_school = None
        count_school = 0
        count_class = 0

        for row in ws.iter_rows(min_row=3, values_only=True):
            number = row[0]
            name = row[1]
            city = row[2]
            principal_name = row[3]
            tel = row[4]
            address = row[5]
            class_name = row[6]
            boys = to_int(row[7])
            girls = to_int(row[8])
            total = to_int(row[9])
            third_grade_total = to_int(row[10])

            if name:
                current_school = JuniorHighSchool.objects.create(
                    number=to_int(number),
                    name=name,
                    city=city or "",
                    principal_name=principal_name or "",
                    tel=tel or "",
                    address=address or "",
                    third_grade_total=third_grade_total,
                )
                count_school += 1

            if current_school and class_name:
                JuniorHighClass.objects.create(
                    school=current_school,
                    class_name=class_name,
                    boys=boys,
                    girls=girls,
                    total=total,
                )
                count_class += 1

        self.stdout.write(self.style.SUCCESS("中学校情報の取込が完了しました"))
        self.stdout.write(f"中学校数: {count_school}")
        self.stdout.write(f"クラス数: {count_class}")