import os
import re

from io import BytesIO
from pathlib import Path
from urllib.parse import quote
from datetime import datetime

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.staticfiles import finders
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.db.models import (
    Q,
    Case,
    When,
    Value,
    IntegerField,
    Count,
    Exists,
    OuterRef,
)
from django.http import HttpResponse, JsonResponse
from django.shortcuts import (
    get_object_or_404,
    redirect,
    render,
)
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph
from reportlab.platypus import Table, TableStyle

from .decorators import (
    active_teacher_required,
    club_advisor_required,
    publicity_admin_required,
    system_admin_required,
)

from .forms import (
    get_seasonal_greeting,
    JuniorHighSchoolForm,
    JuniorHighClassForm,
    ProspectiveStudentForm,
    ScholarshipRequestDocumentForm,
    TeacherPermissionFormSet,
)

from .models import (
    Club,
    DocumentHistory,
    DocumentNumberSequence,
    DormitoryBenefitCategory,
    DormitoryBenefitHistory,
    DormitoryBenefitQuota,
    JuniorHighClass,
    JuniorHighSchool,
    ProspectiveStudent,
    ScholarshipAssignment,
    ScholarshipCategory,
    ScholarshipInterview,
    ScholarshipQuota,
    ScholarshipRankHistory,
    ScholarshipRequestDocument,
    ScholarshipShippingRecord,
    Teacher,
)

def top(request):
    teacher = None

    if request.user.is_authenticated:
        try:
            teacher = request.user.publicity_teacher
        except Teacher.DoesNotExist:
            teacher = None

    # --------------------------------------------------------
    # システムを利用できるか
    # --------------------------------------------------------

    can_use_system = (
        request.user.is_authenticated
        and (
            request.user.is_superuser
            or (
                teacher is not None
                and teacher.is_active
            )
        )
    )

    # --------------------------------------------------------
    # 広報機能を管理できるか
    # --------------------------------------------------------

    can_manage_publicity = (
        request.user.is_superuser
        or (
            teacher is not None
            and teacher.is_active
            and teacher.role in [
                "system_admin",
                "publicity_admin",
            ]
        )
    )

    # --------------------------------------------------------
    # 教員・権限管理を操作できるか
    # --------------------------------------------------------

    can_manage_permissions = (
        request.user.is_superuser
        or (
            teacher is not None
            and teacher.is_active
            and teacher.role == "system_admin"
        )
    )

    # --------------------------------------------------------
    # 奨学生・面談管理を利用できるか
    # --------------------------------------------------------

    can_manage_scholarship = (
        request.user.is_superuser
        or (
            teacher is not None
            and teacher.is_active
            and teacher.role in {
                "club_advisor",
                "publicity_admin",
                "system_admin",
            }
        )
    )

    # --------------------------------------------------------
    # TOP表示
    # --------------------------------------------------------

    return render(
        request,
        "publicity/top.html",
        {
            "teacher": teacher,
            "can_use_system": can_use_system,
            "can_manage_publicity": can_manage_publicity,
            "can_manage_permissions": can_manage_permissions,

            # ★これが必要
            "can_manage_scholarship": can_manage_scholarship,
        },
    )

@active_teacher_required
def school_list(request):
    schools = JuniorHighSchool.objects.prefetch_related("classes").order_by("number", "id")

    return render(request, "publicity/school_list.html", {
        "schools": schools
    })

@club_advisor_required
def school_edit(request, pk):
    school = get_object_or_404(JuniorHighSchool, pk=pk)

    if request.method == "POST":
        form = JuniorHighSchoolForm(request.POST, instance=school)

        if form.is_valid():
            form.save()
            return redirect("publicity:school_list")

    else:
        form = JuniorHighSchoolForm(instance=school)

    return render(request, "publicity/school_form.html", {
        "form": form,
        "school": school,
    })

@publicity_admin_required
def school_delete(request, pk):
    school = get_object_or_404(JuniorHighSchool, pk=pk)
    school.delete()

    return redirect("publicity:school_list")

@active_teacher_required
def school_print(request):
    school_ids = request.GET.getlist("school_ids")

    if school_ids:
        schools = JuniorHighSchool.objects.filter(
            id__in=school_ids
        ).order_by("number", "id")
    else:
        schools = JuniorHighSchool.objects.all().order_by("number", "id")

    return render(request, "publicity/school_print.html", {
        "schools": schools
    })

@publicity_admin_required
def label_select(request):
    schools = JuniorHighSchool.objects.all().order_by("number", "id")

    return render(request, "publicity/label_select.html", {
        "schools": schools
    })

# ============================================================
# 宮崎県内 資料配布管理
# ============================================================

@publicity_admin_required
def miyazaki_material_distribution(request):
    """
    宛名シール・資料印刷の対象校選択画面。

    purpose:
        open_school
            オープンスクール用
            → 宮崎県内の有効な中学校

        prospective
            募集対象生徒登録校
            → ProspectiveStudent が1名以上いる中学校
    """

    # ========================================================
    # GET条件
    # ========================================================

    purpose = (
        request.GET
        .get("purpose", "open_school")
        .strip()
    )

    area = (
        request.GET
        .get("area", "")
        .strip()
    )

    keyword = (
        request.GET
        .get("q", "")
        .strip()
    )

    # ========================================================
    # 印刷目的
    # ========================================================

    if purpose not in {
        "open_school",
        "prospective",
    }:
        purpose = "open_school"

    # ========================================================
    # 基本QuerySet
    # ========================================================

    schools = (
        JuniorHighSchool.objects
        .filter(
            is_active=True,
        )
    )

    # ========================================================
    # 目的別の母集団
    # ========================================================

    if purpose == "open_school":

        # ----------------------------------------------------
        # オープンスクール用
        #
        # 宮崎県内の中学校マスタを対象
        # ----------------------------------------------------

        schools = schools.filter(
            prefecture="宮崎県"
        )

    elif purpose == "prospective":

        # ----------------------------------------------------
        # 募集対象生徒登録校
        #
        # 県内外を問わず、
        # 有効な募集対象生徒が1名以上いる学校
        # ----------------------------------------------------

        schools = (
            schools
            .filter(
                prospective_students__is_active=True
            )
            .distinct()
        )

    # ========================================================
    # 地域条件
    # ========================================================

    main_cities = [
        "小林市",
        "高原町",
        "えびの市",
        "都城市",
    ]

    if area == "local":

        schools = schools.filter(
            city__in=main_cities
        )

    elif area == "other":

        schools = schools.exclude(
            city__in=main_cities
        )

    elif area == "miyazaki":

        schools = schools.filter(
            prefecture="宮崎県"
        )

    elif area == "outside_miyazaki":

        # オープンスクール用では
        # 元々宮崎県内限定なので0件になる。
        schools = schools.exclude(
            prefecture="宮崎県"
        )

    elif area == "kobayashi":

        schools = schools.filter(
            city="小林市"
        )

    elif area == "takaharu":

        schools = schools.filter(
            city="高原町"
        )

    elif area == "ebino":

        schools = schools.filter(
            city="えびの市"
        )

    elif area == "miyakonojo":

        schools = schools.filter(
            city="都城市"
        )

    # ========================================================
    # キーワード検索
    # ========================================================

    if keyword:

        schools = schools.filter(
            Q(
                name__icontains=keyword
            )
            | Q(
                prefecture__icontains=keyword
            )
            | Q(
                city__icontains=keyword
            )
            | Q(
                address__icontains=keyword
            )
            | Q(
                official_address__icontains=keyword
            )
        )

    # ========================================================
    # 地域優先順
    # ========================================================

    schools = (
        schools
        .annotate(
            area_order=Case(

                When(
                    city__icontains="小林",
                    then=Value(1),
                ),

                When(
                    city__icontains="高原",
                    then=Value(2),
                ),

                When(
                    city__icontains="えびの",
                    then=Value(3),
                ),

                When(
                    city__icontains="都城",
                    then=Value(4),
                ),

                When(
                    prefecture="宮崎県",
                    then=Value(50),
                ),

                default=Value(99),

                output_field=IntegerField(),
            )
        )
        .order_by(
            "area_order",
            "prefecture",
            "city",
            "name",
        )
    )

    # ========================================================
    # 募集対象生徒数
    # ========================================================

    prospective_counts = {

        row[
            "junior_high_school_id"
        ]: row[
            "student_count"
        ]

        for row in (
            ProspectiveStudent.objects
            .filter(
                is_active=True
            )
            .values(
                "junior_high_school_id"
            )
            .annotate(
                student_count=Count("id")
            )
        )
    }

    # ========================================================
    # Template用
    # ========================================================

    school_list = []

    for school in schools:

        school.prospective_count = (
            prospective_counts.get(
                school.id,
                0,
            )
        )

        school_list.append(
            school
        )

    # ========================================================
    # Context
    # ========================================================

    context = {

        "schools": school_list,

        "selected_purpose": purpose,

        "selected_area": area,

        "keyword": keyword,

        "school_count": len(
            school_list
        ),

    }

    return render(
        request,
        (
            "publicity/"
            "miyazaki_material_distribution.html"
        ),
        context,
    )

@publicity_admin_required
def label_pdf(request):

    # ========================================================
    # GETパラメータ
    # ========================================================

    school_ids = request.GET.getlist(
        "school_ids"
    )

    # TELを表示するか
    # show_tel=1 の場合のみ表示
    show_tel = (
        request.GET.get(
            "show_tel",
            "0"
        )
        == "1"
    )

    # ========================================================
    # 対象中学校
    # ========================================================

    if school_ids:

        schools = (
            JuniorHighSchool.objects
            .filter(
                id__in=school_ids
            )
            .order_by(
                "number",
                "id",
            )
        )

    else:

        schools = (
            JuniorHighSchool.objects
            .all()
            .order_by(
                "number",
                "id",
            )
        )

    # ========================================================
    # PDFレスポンス
    # ========================================================

    response = HttpResponse(
        content_type="application/pdf"
    )

    response[
        "Content-Disposition"
    ] = (
        'inline; '
        'filename="junior_high_labels.pdf"'
    )

    p = canvas.Canvas(
        response,
        pagesize=A4,
    )

    # ========================================================
    # フォント
    # ========================================================

    font_path = os.path.join(
        settings.BASE_DIR,
        "static",
        "fonts",
        "ipaexm.ttf",
    )

    pdfmetrics.registerFont(
        TTFont(
            "IPAexMincho",
            font_path,
        )
    )

    # ========================================================
    # 用紙設定
    #
    # FJA210A
    # A4 / 12面 / 2列×6段
    # ========================================================

    page_width, page_height = A4

    label_width = 83.8 * mm
    label_height = 42.3 * mm

    margin_left = 21.2 * mm
    margin_top = 21.2 * mm

    cols = 2
    labels_per_page = 12

    # ========================================================
    # 右列位置補正
    #
    # 担当者要望：
    # 右側の列のみ左へ5mm移動
    # ========================================================

    right_column_adjustment = 5 * mm

    label_index = 0

    # ========================================================
    # ラベル生成
    # ========================================================

    for school in schools:

        pos = (
            label_index
            % labels_per_page
        )

        col = pos % cols
        row = pos // cols

        # ----------------------------------------------------
        # ラベル基準位置
        # ----------------------------------------------------

        x = (
            margin_left
            + (
                col
                * label_width
            )
        )

        # 右列だけ左へ5mm
        if col == 1:

            x += right_column_adjustment

        y = (
            page_height
            - margin_top
            - (
                (row + 1)
                * label_height
            )
        )

        # ----------------------------------------------------
        # ラベル内余白
        # ----------------------------------------------------

        text_x = (
            x
            + 5 * mm
        )

        text_y = (
            y
            + label_height
            - 7 * mm
        )

        # ====================================================
        # 郵便番号
        # ====================================================

        postal_code = (
            school.official_postal_code
            or ""
        ).strip()

        # 念のため「〒」がDB側に入っていても二重にしない
        postal_code = (
            postal_code
            .replace("〒", "")
            .strip()
        )

        if postal_code:

            postal_text = (
                f"〒{postal_code}"
            )

        else:

            postal_text = ""

        p.setFont(
            "IPAexMincho",
            8,
        )

        p.drawString(
            text_x,
            text_y,
            postal_text,
        )

        # ====================================================
        # 住所
        #
        # 文科省データの正式住所を優先。
        # なければ従来address。
        # ====================================================

        address = (
            school.official_address
            or school.address
            or ""
        ).strip()

        p.setFont(
            "IPAexMincho",
            8,
        )

        p.drawString(
            text_x,
            text_y - 6 * mm,
            address,
        )

        # ====================================================
        # 学校名
        # ====================================================

        p.setFont(
            "IPAexMincho",
            10,
        )

        p.drawString(
            text_x,
            text_y - 15 * mm,
            school.name,
        )

        # ====================================================
        # 宛名
        # ====================================================

        p.setFont(
            "IPAexMincho",
            9,
        )

        p.drawString(
            text_x,
            text_y - 23 * mm,
            "学校長　様",
        )

        # ====================================================
        # TEL
        #
        # show_tel=1 の場合のみ表示
        # ====================================================

        if (
            show_tel
            and school.tel
        ):

            p.setFont(
                "IPAexMincho",
                7,
            )

            p.drawString(
                text_x,
                text_y - 30 * mm,
                f"TEL：{school.tel}",
            )

        # ====================================================
        # 次のラベル
        # ====================================================

        label_index += 1

        # 12枚ごとに改ページ
        if (
            label_index
            % labels_per_page
            == 0
        ):

            p.showPage()

    # ========================================================
    # PDF終了
    # ========================================================

    p.save()

    return response

@publicity_admin_required
def class_label_pdf(request):
    school_ids = request.GET.getlist("school_ids")

    if school_ids:
        classes = JuniorHighClass.objects.select_related(
            "school"
        ).filter(
            school_id__in=school_ids
        ).order_by("school__number", "school__name", "class_name")
    else:
        classes = JuniorHighClass.objects.select_related(
            "school"
        ).order_by("school__number", "school__name", "class_name")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="class_labels.pdf"'

    p = canvas.Canvas(response, pagesize=A4)

    font_path = os.path.join(
        settings.BASE_DIR,
        "static",
        "fonts",
        "ipaexm.ttf"
    )

    pdfmetrics.registerFont(TTFont("IPA", font_path))

    p.setFont("IPA", 10)

    width, height = A4

    label_width = 83.8 * mm
    label_height = 42.3 * mm

    margin_left = 21.2 * mm
    margin_top = 21.2 * mm

    cols = 2
    labels_per_page = 12

    label_index = 0

    for cls in classes:
        pos = label_index % labels_per_page
        col = pos % cols
        row = pos // cols

        x = margin_left + col * label_width
        y = height - margin_top - (row + 1) * label_height

        p.roundRect(
            x + 3 * mm,
            y + 5 * mm,
            label_width - 6 * mm,
            label_height - 10 * mm,
            3 * mm
        )

        p.setFont("IPA", 10)
        p.drawCentredString(
            x + label_width / 2,
            y + 25 * mm,
            cls.school.name
        )

        p.setFont("IPA", 10)
        p.drawCentredString(
            x + label_width / 2,
            y + 15 * mm,
            f"{cls.class_name}　{cls.total}名"
        )

        label_index += 1

        if label_index % labels_per_page == 0:
            p.showPage()
            p.setFont("IPA", 10)

    p.save()

    return response

@club_advisor_required
def class_edit(request, pk):
    school_class = get_object_or_404(JuniorHighClass, pk=pk)

    if request.method == "POST":
        form = JuniorHighClassForm(request.POST, instance=school_class)

        if form.is_valid():
            form.save()
            return redirect("publicity:school_list")

    else:
        form = JuniorHighClassForm(instance=school_class)

    return render(request, "publicity/class_form.html", {
        "form": form,
        "school_class": school_class,
    })

@system_admin_required
def teacher_permission_manage(request):
    queryset = Teacher.objects.all().order_by(
        "employee_number"
    )

    if request.method == "POST":
        formset = TeacherPermissionFormSet(
            request.POST,
            queryset=queryset,
        )

        if formset.is_valid():
            changed_teachers = formset.save()

            messages.success(
                request,
                f"{len(changed_teachers)}件の権限情報を更新しました。",
            )

            return redirect(
                "publicity:teacher_permission_manage"
            )

    else:
        formset = TeacherPermissionFormSet(
            queryset=queryset
        )

    teacher_rows = zip(
        queryset,
        formset.forms,
    )

    return render(
        request,
        "publicity/teacher_permission_manage.html",
        {
            "formset": formset,
            "teacher_rows": teacher_rows,
        },
    )

def login_denied(request):
    return render(
        request,
        "publicity/login_denied.html",
        status=403,
    )

@club_advisor_required
def prospective_student_list(request):
    teacher = request.teacher

    students = (
        ProspectiveStudent.objects
        .select_related(
            "junior_high_school",
            "club",
            "scholarship_category",
            "registered_by",
            "assigned_teacher",
        )
        .filter(is_active=True)
    )

    # 広報管理者・システム管理者は全件表示
    if teacher.role not in {
        "publicity_admin",
        "system_admin",
    }:
        students = students.filter(
            Q(registered_by=teacher)
            | Q(assigned_teacher=teacher)
        )

    keyword = request.GET.get("q", "").strip()
    school_id = request.GET.get("school", "").strip()
    club_id = request.GET.get("club", "").strip()

    if keyword:
        students = students.filter(
            Q(name__icontains=keyword)
            | Q(junior_high_school__name__icontains=keyword)
            | Q(club__name__icontains=keyword)
        )

    if school_id.isdigit():
        students = students.filter(
            junior_high_school_id=school_id
        )

    if club_id.isdigit():
        students = students.filter(
            club_id=club_id
        )

    students = students.order_by(
        "-created_at",
        "name",
    )

    visible_school_ids = (
        students
        .values_list(
            "junior_high_school_id",
            flat=True,
        )
        .distinct()
    )

    visible_club_ids = (
        students
        .values_list(
            "club_id",
            flat=True,
        )
        .distinct()
    )

    schools = (
        JuniorHighSchool.objects
        .filter(id__in=visible_school_ids)
        .order_by("city", "name")
    )

    clubs = (
        Club.objects
        .filter(id__in=visible_club_ids)
        .order_by("name")
    )

    return render(
        request,
        "publicity/prospective_student_list.html",
        {
            "students": students,
            "schools": schools,
            "clubs": clubs,
            "keyword": keyword,
            "selected_school": school_id,
            "selected_club": club_id,
        },
    )

@club_advisor_required
def prospective_student_excel(request):

    teacher = request.teacher

    # ============================================================
    # 基本QuerySet
    # ============================================================

    students = (
        ProspectiveStudent.objects
        .select_related(
            "junior_high_school",
            "club",
            "registered_by",
            "assigned_teacher",
        )
        .filter(
            is_active=True
        )
    )

    # ============================================================
    # 権限制御
    #
    # 広報管理者・システム管理者は全件
    # それ以外は自分に関係する生徒のみ
    # ============================================================

    if teacher.role not in {
        "publicity_admin",
        "system_admin",
    }:

        students = students.filter(
            Q(
                registered_by=teacher
            )
            | Q(
                assigned_teacher=teacher
            )
        )

    # ============================================================
    # 一覧画面と同じ検索条件
    # ============================================================

    keyword = request.GET.get(
        "q",
        "",
    ).strip()

    school_id = request.GET.get(
        "school",
        "",
    ).strip()

    club_id = request.GET.get(
        "club",
        "",
    ).strip()

    if keyword:

        students = students.filter(
            Q(
                name__icontains=keyword
            )
            | Q(
                name_kana__icontains=keyword
            )
            | Q(
                junior_high_school__name__icontains=keyword
            )
            | Q(
                club__name__icontains=keyword
            )
        )

    if school_id.isdigit():

        students = students.filter(
            junior_high_school_id=school_id
        )

    if club_id.isdigit():

        students = students.filter(
            club_id=club_id
        )

    # ============================================================
    # 地域順
    #
    # 小林 → えびの → 高原 → 都城 → 三股 → 宮崎
    # その他は後ろ
    # ============================================================

    students = (
        students
        .annotate(
            area_order=Case(

                When(
                    junior_high_school__city__icontains="小林",
                    then=Value(1),
                ),

                When(
                    junior_high_school__city__icontains="えびの",
                    then=Value(2),
                ),

                When(
                    junior_high_school__city__icontains="高原",
                    then=Value(3),
                ),

                When(
                    junior_high_school__city__icontains="都城",
                    then=Value(4),
                ),

                When(
                    junior_high_school__city__icontains="三股",
                    then=Value(5),
                ),

                When(
                    junior_high_school__city__icontains="宮崎",
                    then=Value(6),
                ),

                # その他宮崎県内
                When(
                    junior_high_school__prefecture__icontains="宮崎",
                    then=Value(90),
                ),

                # 県外
                default=Value(99),

                output_field=IntegerField(),
            )
        )
        .order_by(
            "area_order",
            "junior_high_school__city",
            "junior_high_school__name",
            "club__name",
            "name_kana",
            "name",
        )
    )

    # ============================================================
    # Excel作成
    # ============================================================

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = (
        "管理職用募集対象生徒一覧"
    )

    # ============================================================
    # タイトル
    # ============================================================

    worksheet.merge_cells(
        "A1:K1"
    )

    title_cell = worksheet["A1"]

    title_cell.value = (
        "募集対象生徒 管理職確認・連絡メモ"
    )

    title_cell.font = Font(
        bold=True,
        size=16,
        color="FFFFFF",
    )

    title_cell.fill = PatternFill(
        fill_type="solid",
        fgColor="28556B",
    )

    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.row_dimensions[1].height = 30

    # ============================================================
    # 説明
    # ============================================================

    worksheet.merge_cells(
        "A2:K2"
    )

    description_cell = (
        worksheet["A2"]
    )

    description_cell.value = (
        "中学校からの連絡内容、"
        "部顧問への確認結果、面談予定等を"
        "自由に記録してください。"
    )

    description_cell.alignment = Alignment(
        vertical="center",
        wrap_text=True,
    )

    worksheet.row_dimensions[2].height = 28

    # ============================================================
    # 見出し
    # ============================================================

    headers = [
        "No.",
        "地域",
        "中学校",
        "部活動",
        "生徒氏名",
        "ふりがな",
        "担当教員",
        "中学校からの連絡",
        "部顧問返答",
        "面談",
        "管理職メモ",
    ]

    worksheet.append(
        headers
    )

    header_row = 3

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="DDEBF2",
    )

    header_font = Font(
        bold=True,
        size=12.5,
    )

    for cell in worksheet[
        header_row
    ]:

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.row_dimensions[
        header_row
    ].height = 28

    # ============================================================
    # 生徒データ
    # ============================================================

    for index, student in enumerate(
        students,
        start=1,
    ):

        school = (
            student.junior_high_school
        )

        area_name = ""

        if school:

            area_name = (
                school.city
                or school.prefecture
                or ""
            )

        worksheet.append(
            [
                index,

                area_name,

                (
                    school.name
                    if school
                    else ""
                ),

                (
                    student.club.name
                    if student.club
                    else ""
                ),

                student.name,

                (
                    student.name_kana
                    or ""
                ),

                (
                    student.assigned_teacher.name
                    if student.assigned_teacher
                    else ""
                ),

                # 中学校からの連絡
                "",

                # 部顧問返答
                "",

                # 面談
                "",

                # 管理職メモ
                "",
            ]
        )

    # ============================================================
    # 列幅
    # ============================================================

    column_widths = {

        "A": 6,   # No.

        "B": 13,  # 地域

        "C": 28,  # 中学校

        "D": 18,  # 部活動

        "E": 16,  # 生徒氏名

        "F": 18,  # ふりがな

        "G": 15,  # 担当教員

        "H": 22,  # 中学校からの連絡

        "I": 16,  # 部顧問返答

        "J": 15,  # 面談

        "K": 40,  # 管理職メモ
    }

    for column, width in (
        column_widths.items()
    ):

        worksheet.column_dimensions[
            column
        ].width = width

    # ============================================================
    # データ行装飾
    # ============================================================

    thin_border = Border(
        left=Side(
            style="thin",
            color="B8C6CC",
        ),
        right=Side(
            style="thin",
            color="B8C6CC",
        ),
        top=Side(
            style="thin",
            color="B8C6CC",
        ),
        bottom=Side(
            style="thin",
            color="B8C6CC",
        ),
    )

    for row in worksheet.iter_rows(
        min_row=4,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=11,
    ):

        for cell in row:

            cell.border = thin_border

            # 文字サイズを約1.5ptアップ
            cell.font = Font(
                size=12.5,
            )

            # 基本は中央寄せ
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

    # ============================================================
    # 管理職メモ欄だけ左寄せ
    # ============================================================

    for row_number in range(
        4,
        worksheet.max_row + 1,
    ):

        worksheet[
            f"K{row_number}"
        ].alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )

    # ============================================================
    # 手書きしやすいようにデータ行を高くする
    # ============================================================

    for row_number in range(
        4,
        worksheet.max_row + 1,
    ):

        worksheet.row_dimensions[
            row_number
        ].height = 42

    # ============================================================
    # 学校が切り替わる位置を強調
    # ============================================================

    previous_school = None

    for row_number in range(
        4,
        worksheet.max_row + 1,
    ):

        current_school = (
            worksheet[
                f"C{row_number}"
            ].value
        )

        if (
            previous_school is not None
            and current_school
            != previous_school
        ):

            for cell in worksheet[
                row_number
            ]:

                cell.border = Border(
                    left=Side(
                        style="thin",
                        color="B8C6CC",
                    ),
                    right=Side(
                        style="thin",
                        color="B8C6CC",
                    ),
                    top=Side(
                        style="medium",
                        color="28556B",
                    ),
                    bottom=Side(
                        style="thin",
                        color="B8C6CC",
                    ),
                )

        previous_school = (
            current_school
        )

    # ============================================================
    # 固定
    # ============================================================

    worksheet.freeze_panes = (
        "A4"
    )

    # ============================================================
    # オートフィルター
    # ============================================================

    if worksheet.max_row >= 3:

        worksheet.auto_filter.ref = (
            f"A3:K{worksheet.max_row}"
        )

    # ============================================================
    # 印刷設定
    # ============================================================

    worksheet.sheet_properties.pageSetUpPr.fitToPage = (
        True
    )

    worksheet.page_setup.fitToWidth = (
        1
    )

    worksheet.page_setup.fitToHeight = (
        0
    )

    worksheet.page_setup.orientation = (
        "landscape"
    )

    worksheet.page_setup.paperSize = (
        worksheet.PAPERSIZE_A4
    )

    worksheet.print_title_rows = (
        "1:3"
    )

    worksheet.sheet_view.showGridLines = (
        False
    )

    # ============================================================
    # 印刷余白
    # ============================================================

    worksheet.page_margins.left = (
        0.25
    )

    worksheet.page_margins.right = (
        0.25
    )

    worksheet.page_margins.top = (
        0.4
    )

    worksheet.page_margins.bottom = (
        0.4
    )

    # ============================================================
    # 保存
    # ============================================================

    output = BytesIO()

    workbook.save(
        output
    )

    output.seek(0)

    filename = (
        "募集対象生徒_管理職確認メモ.xlsx"
    )

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
    )

    response[
        "Content-Disposition"
    ] = (
        "attachment; "
        "filename*=UTF-8''"
        f"{quote(filename)}"
    )

    return response

def _get_school_principals():
    return {
        str(school.id): school.principal_name or "未登録"
        for school in JuniorHighSchool.objects.all().only(
            "id",
            "principal_name",
        )
    }

@club_advisor_required
def prospective_student_create(request):
    teacher = request.teacher

    if request.method == "POST":
        form = ProspectiveStudentForm(
            request.POST
        )

        if form.is_valid():
            student = form.save(commit=False)

            # 登録者・担当者はログイン中の先生
            student.registered_by = teacher
            student.assigned_teacher = teacher
            student.is_active = True

            student.save()

            messages.success(
                request,
                f"{student.name}さんを登録しました。",
            )

            return redirect(
                "publicity:prospective_student_list"
            )

    else:
        form = ProspectiveStudentForm()

    return render(
        request,
        "publicity/prospective_student_form.html",
        {
            "form": form,
            "page_title": "募集対象生徒を登録",
            "submit_label": "登録する",
            "school_principals": _get_school_principals(),
            "is_edit": False,
        },
    )

@club_advisor_required
def prospective_student_edit(request, pk):
    teacher = request.teacher

    queryset = (
        ProspectiveStudent.objects
        .select_related(
            "junior_high_school",
            "club",
            "registered_by",
            "assigned_teacher",
        )
        .filter(is_active=True)
    )

    if teacher.role not in {
        "publicity_admin",
        "system_admin",
    }:
        queryset = queryset.filter(
            Q(registered_by=teacher)
            | Q(assigned_teacher=teacher)
        )

    student = get_object_or_404(
        queryset,
        pk=pk,
    )

    if request.method == "POST":
        form = ProspectiveStudentForm(
            request.POST,
            instance=student,
        )

        if form.is_valid():
            form.save()

            messages.success(
                request,
                f"{student.name}さんの情報を更新しました。",
            )

            return redirect(
                "publicity:prospective_student_list"
            )

    else:
        form = ProspectiveStudentForm(
            instance=student
        )

    return render(
        request,
        "publicity/prospective_student_form.html",
        {
            "form": form,
            "student": student,
            "page_title": "募集対象生徒を編集",
            "submit_label": "変更を保存",
            "school_principals": _get_school_principals(),
            "is_edit": True,
        },
    )

@club_advisor_required
def junior_high_school_search(request):
    keyword = (
        request.GET
        .get("q", "")
        .strip()
    )

    prefecture = (
        request.GET
        .get("prefecture", "")
        .strip()
    )

    schools = (
        JuniorHighSchool.objects
        .filter(is_active=True)
        .order_by(
            "prefecture",
            "city",
            "name",
        )
    )

    if prefecture:
        schools = schools.filter(
            prefecture=prefecture
        )

    if keyword:
        # 全角スペースを半角スペースへ統一し、
        # スペース区切りで複数の検索語に分割
        search_words = (
            keyword
            .replace("　", " ")
            .split()
        )

        # 各検索語についてfilterを重ねることで
        # AND検索にする
        for word in search_words:
            schools = schools.filter(
                Q(name__icontains=word)
                | Q(prefecture__icontains=word)
                | Q(city__icontains=word)
                | Q(address__icontains=word)
                | Q(
                    official_address__icontains=word
                )
                | Q(
                    official_postal_code__icontains=word
                )
                | Q(school_code__icontains=word)
            )

    else:
        # 未入力時に全国の学校を返さない
        schools = schools.none()

    # 候補を最大30校に制限
    schools = schools[:30]

    results = [
        {
            "id": school.id,
            "name": school.name,
            "prefecture": (
                school.prefecture or ""
            ),
            "city": school.city or "",
            "principal_name": (
                school.principal_name or ""
            ),
            "postal_code": (
                school.official_postal_code
                or ""
            ),
            "address": (
                school.address
                or school.official_address
                or ""
            ),
            "label": (
                f"{school.name}"
                f"（{school.prefecture or ''}"
                f"{school.city or ''}）"
            ),
        }
        for school in schools
    ]

    return JsonResponse(
        {
            "results": results,
        },
        json_dumps_params={
            "ensure_ascii": False,
        },
    )

@system_admin_required
def document_management_top(request):
    return render(
        request,
        "publicity/document_management_top.html",
    )

@system_admin_required
def scholarship_document_student_select(request):

    # ========================================================
    # GET条件
    # ========================================================

    keyword = request.GET.get(
        "q",
        "",
    ).strip()

    club_id = request.GET.get(
        "club",
        "",
    ).strip()

    region = request.GET.get(
        "region",
        "",
    ).strip()

    area = request.GET.get(
        "area",
        "",
    ).strip()

    # 複数都道府県
    selected_prefectures = [
        prefecture
        for prefecture in request.GET.getlist("prefecture")
        if prefecture
    ]

    # 複数中学校
    selected_school_ids = [
        school_id
        for school_id in request.GET.getlist("school")
        if school_id.isdigit()
    ]

    # ========================================================
    # 基本QuerySet
    # ========================================================

    students = (
        ProspectiveStudent.objects
        .filter(
            is_active=True
        )
        .select_related(
            "junior_high_school",
            "club",
            "assigned_teacher",
        )
    )

    # ========================================================
    # キーワード検索
    # ========================================================

    if keyword:

        students = students.filter(
            Q(name__icontains=keyword)
            | Q(name_kana__icontains=keyword)
            | Q(
                junior_high_school__name__icontains=keyword
            )
            | Q(
                club__name__icontains=keyword
            )
        )

    # ========================================================
    # 部活動
    # ========================================================

    if club_id.isdigit():

        students = students.filter(
            club_id=club_id
        )

    # ========================================================
    # 地域
    # ========================================================

    direct_cities = [
        "小林市",
        "高原町",
        "えびの市",
        "都城市",
    ]

    if region == "miyazaki":

        students = students.filter(
            junior_high_school__prefecture="宮崎県"
        )

        # 宮崎県内エリア
        if area == "direct":

            students = students.filter(
                junior_high_school__city__in=direct_cities
            )

        elif area == "other":

            students = students.exclude(
                junior_high_school__city__in=direct_cities
            )

        # 県外用選択は無効化
        selected_prefectures = []

    elif region == "outside":

        students = students.exclude(
            junior_high_school__prefecture="宮崎県"
        )

        # 選択された都道府県があれば絞り込む
        if selected_prefectures:

            students = students.filter(
                junior_high_school__prefecture__in=selected_prefectures
            )

        # 県内エリアは無効
        area = ""

    else:

        # 地域「すべて」
        area = ""
        selected_prefectures = []

    # ========================================================
    # この時点で選択可能な中学校を取得
    #
    # 学校フィルタを掛ける「前」に作るのが重要
    # ========================================================

    school_source_students = students

    schools = (
        JuniorHighSchool.objects
        .filter(
            id__in=school_source_students.values_list(
                "junior_high_school_id",
                flat=True,
            )
        )
        .order_by(
            "prefecture",
            "city",
            "name",
        )
        .distinct()
    )

    # ========================================================
    # 中学校 複数選択
    # ========================================================

    if selected_school_ids:

        students = students.filter(
            junior_high_school_id__in=selected_school_ids
        )

    # ========================================================
    # 発送済み判定
    # ========================================================

    shipped_records = (
        ScholarshipShippingRecord.objects
        .filter(
            student_id=OuterRef("pk"),
            is_shipped=True,
        )
    )

    students = students.annotate(
        has_been_shipped=Exists(
            shipped_records
        )
    )

    # ========================================================
    # 表示順
    #
    # 小林 → 高原 → えびの → 都城 → その他
    # ========================================================

    students = (
        students
        .annotate(
            area_order=Case(

                When(
                    junior_high_school__city="小林市",
                    then=Value(1),
                ),

                When(
                    junior_high_school__city="高原町",
                    then=Value(2),
                ),

                When(
                    junior_high_school__city="えびの市",
                    then=Value(3),
                ),

                When(
                    junior_high_school__city="都城市",
                    then=Value(4),
                ),

                default=Value(99),

                output_field=IntegerField(),
            )
        )
        .order_by(
            "area_order",
            "junior_high_school__prefecture",
            "junior_high_school__city",
            "junior_high_school__name",
            "name",
        )
    )

    # ========================================================
    # 県外の都道府県一覧
    #
    # 全国47都道府県ではなく、
    # 実際に募集対象生徒がいる県だけ
    # ========================================================

    outside_prefectures = (
        ProspectiveStudent.objects
        .filter(
            is_active=True
        )
        .exclude(
            junior_high_school__prefecture="宮崎県"
        )
        .exclude(
            junior_high_school__prefecture=""
        )
        .exclude(
            junior_high_school__prefecture__isnull=True
        )
        .values(
            "junior_high_school__prefecture"
        )
        .annotate(
            student_count=Count("id")
        )
        .order_by(
            "junior_high_school__prefecture"
        )
    )

    # ========================================================
    # 部活動一覧
    # ========================================================

    clubs = (
        Club.objects
        .filter(
            id__in=ProspectiveStudent.objects
            .filter(
                is_active=True
            )
            .values_list(
                "club_id",
                flat=True,
            )
        )
        .order_by("name")
        .distinct()
    )

    # ========================================================
    # Template
    # ========================================================

    return render(
        request,
        "publicity/scholarship_document_student_select.html",
        {
            "students": students,
            "schools": schools,
            "clubs": clubs,

            "keyword": keyword,

            "selected_club": club_id,

            "selected_region": region,
            "selected_area": area,

            "outside_prefectures": outside_prefectures,
            "selected_prefectures": selected_prefectures,

            "selected_school_ids": selected_school_ids,
        },
    )

@login_required
def scholarship_request_document_create(request):
    """
    奨学生募集依頼文書の作成画面
    """

    today = timezone.localdate()

    # -----------------------------
    # 現在年度を取得
    # -----------------------------
    fiscal_year = get_fiscal_year_from_date(
        today
    )

    # -----------------------------
    # 次回文書番号を確認
    # ※ここではまだ正式採番しない
    # -----------------------------
    sequence = (
        DocumentNumberSequence.objects
        .filter(
            fiscal_year=fiscal_year
        )
        .first()
    )

    if sequence:
        next_number = (
            sequence.last_number + 1
        )
    else:
        next_number = 1

    next_document_number = (
        f"小林西発第"
        f"{next_number:03d}"
        f"号"
    )

    # -----------------------------
    # 選択された中学校
    # -----------------------------
    school_id = (
        request.GET
        .get("school", "")
        .strip()
    )

    # -----------------------------
    # Form
    #
    # request.GETを直接渡さない。
    # schoolだけinitialとして設定する。
    # これで年度・日付・季語・文書番号の
    # initialが正常に表示される。
    # -----------------------------
    initial_data = {}

    if school_id.isdigit():
        initial_data["school"] = (
            school_id
        )

    form = ScholarshipRequestDocumentForm(
        next_document_number=(
            next_document_number
        ),
        initial=initial_data,
    )

    selected_school = None

    students = (
        ProspectiveStudent.objects.none()
    )

    # -----------------------------
    # 中学校・対象生徒取得
    # -----------------------------
    if school_id.isdigit():

        selected_school = (
            form.fields["school"]
            .queryset
            .filter(
                pk=school_id
            )
            .first()
        )

        if selected_school:

            students = (
                ProspectiveStudent.objects
                .filter(
                    junior_high_school=(
                        selected_school
                    ),
                    is_active=True,
                )
                .select_related(
                    "club",
                    "junior_high_school",
                )
                .order_by(
                    "club__name",
                    "name_kana",
                    "name",
                )
            )

    context = {
        "form": form,
        "selected_school": (
            selected_school
        ),
        "students": students,
        "next_document_number": (
            next_document_number
        ),
    }

    return render(
        request,
        (
            "publicity/"
            "scholarship_request_document_create.html"
        ),
        context,
    )

def get_fiscal_year_from_date(date_obj):
    """
    発行日から年度（西暦）を取得する。
    4月～翌年3月を同一年度として扱う。
    """

    if date_obj.month >= 4:
        return date_obj.year

    return date_obj.year - 1


@transaction.atomic
def issue_document_number(issue_date):
    """
    文書番号を正式発行する。

    PDF生成時など、
    本当に文書を発行するときだけ呼び出す。
    """

    fiscal_year = get_fiscal_year_from_date(
        issue_date
    )

    sequence, created = (
        DocumentNumberSequence.objects
        .select_for_update()
        .get_or_create(
            fiscal_year=fiscal_year,
            defaults={
                "last_number": 0,
            },
        )
    )

    sequence.last_number += 1

    sequence.save(
        update_fields=[
            "last_number",
            "updated_at",
        ]
    )

    document_number = (
        f"小林西発第"
        f"{sequence.last_number:03d}"
        f"号"
    )

    return document_number

@system_admin_required
def scholarship_document_confirm(request):

    if request.method != "POST":
        return redirect(
            "publicity:scholarship_document_student_select"
        )

    student_ids = request.POST.getlist(
        "student_ids"
    )

    if not student_ids:
        messages.error(
            request,
            "文書を発行する生徒を選択してください。",
        )

        return redirect(
            "publicity:scholarship_document_student_select"
        )

    students = (
        ProspectiveStudent.objects
        .filter(
            id__in=student_ids,
            is_active=True,
        )
        .select_related(
            "junior_high_school",
            "club",
            "assigned_teacher",
        )
        .order_by(
            "junior_high_school__prefecture",
            "junior_high_school__city",
            "junior_high_school__name",
            "club__name",
            "name",
        )
    )

    # --------------------------------
    # 中学校単位でグループ化
    # --------------------------------

    grouped_schools = {}

    for student in students:

        school = (
            student.junior_high_school
        )

        if school.id not in grouped_schools:

            grouped_schools[
                school.id
            ] = {
                "school": school,
                "students": [],
            }

        grouped_schools[
            school.id
        ]["students"].append(
            student
        )

    school_groups = list(
        grouped_schools.values()
    )

    today = timezone.localdate()

    fiscal_year = (
        get_fiscal_year_from_date(
            today
        )
    )

    sequence = (
        DocumentNumberSequence.objects
        .filter(
            fiscal_year=fiscal_year
        )
        .first()
    )

    if sequence:
        last_number = sequence.last_number
    else:
        last_number = 0

    for index, group in enumerate(
        school_groups,
        start=1,
    ):
        group[
            "suggested_document_number"
        ] = (
            f"小林西発第"
            f"{last_number + index:03d}"
            f"号"
        )

    reiwa_year = (
        fiscal_year - 2018
    )
    # --------------------------------
    # 募集年度
    # 現在年度 + 1年度を初期値にする
    # --------------------------------

    recruitment_fiscal_year = (
        fiscal_year + 1
    )

    recruitment_reiwa_year = (
        recruitment_fiscal_year - 2018
    )

    issue_reiwa_year = (
        today.year - 2018
    )

    issue_date_label = (
        f"令和{issue_reiwa_year}年"
        f"{today.month}月"
        f"{today.day}日"
    )

    seasonal_greeting = (
        get_seasonal_greeting(
            today.month
        )
    )

    return render(
        request,
        "publicity/scholarship_document_confirm.html",
        {
            "students": students,
            "school_groups": school_groups,
            "student_ids": student_ids,
            "student_count": students.count(),
            "school_count": len(
                school_groups
            ),
            "fiscal_year_label": (
                f"令和{recruitment_reiwa_year}年度"
            ),
            "issue_date": today,
            "issue_date_label": (
                issue_date_label
            ),
            "seasonal_greeting": (
                seasonal_greeting
            ),
        },
    )

def draw_scholarship_request_page(
    pdf,
    school,
    students,
    document_number,
    issue_date,
    fiscal_year,
    recruitment_year,
    seasonal_greeting,
    principal_name,
):
    """
    高千穂学園奨学生募集依頼文書を
    PDFの1ページとして描画する。

    1～10名：
        標準レイアウト

    11～20名：
        コンパクトレイアウト

    fiscal_year:
        文書管理上の年度

    recruitment_year:
        奨学生の募集年度
        例：「令和9年度」

    ※ showPage() / save() はここでは行わない。
    """

    width, height = A4

    # ============================================================
    # 生徒数
    # ============================================================

    student_list = list(students)

    student_count = len(
        student_list
    )

    # ============================================================
    # レイアウトモード
    # ============================================================

    compact_mode = (
        student_count > 10
    )

    # ============================================================
    # フォント
    # ============================================================

    font_min = "HeiseiMin-W3"
    font_gothic = "HeiseiKakuGo-W5"

    pdfmetrics.registerFont(
        UnicodeCIDFont(
            font_min
        )
    )

    pdfmetrics.registerFont(
        UnicodeCIDFont(
            font_gothic
        )
    )

    # 日本語1文字分の字下げ
    em = 11

    # ============================================================
    # 人数に応じたレイアウト設定
    # ============================================================

    if compact_mode:

        # --------------------------------------------------------
        # 11～16名
        # 標準版に近い読みやすさを維持
        # --------------------------------------------------------

        title_y = (
            height - 103 * mm
        )

        title_font_size = 12.2

        body_font_size = 10.0

        body_leading = 14

        body_first_indent = 10.0

        body_start_gap = 9 * mm

        paragraph_gap = 1.5 * mm

        # 本文終了後、「記」まで少しだけ空ける
        before_ki_gap = 4 * mm

        # 「記」から表まで
        after_ki_gap = 6 * mm

        table_font_size = 9.3

        table_leading = 10.5

        table_top_padding = 1.7

        table_bottom_padding = 1.7

        table_left_padding = 4.5

        table_right_padding = 4.5

        final_gap = 5 * mm

        final_font_size = 10.5

    else:

        # --------------------------------------------------------
        # 1～10名
        # --------------------------------------------------------

        title_y = (
            height - 108 * mm
        )

        title_font_size = 12.5

        body_font_size = 10.5

        body_leading = 16

        body_first_indent = 10.5

        body_start_gap = 11 * mm

        paragraph_gap = 2.5 * mm

        before_ki_gap = 2 * mm

        after_ki_gap = 8 * mm

        table_font_size = 9.5

        table_leading = 11

        table_top_padding = 3

        table_bottom_padding = 3

        table_left_padding = 5

        table_right_padding = 5

        final_gap = 7 * mm

        final_font_size = 10.5

    # ============================================================
    # 右上：文書番号・発行日
    # ============================================================

    pdf.setFont(
        font_min,
        10.5,
    )

    pdf.drawRightString(
        width - 25 * mm,
        height - 25 * mm,
        document_number,
    )

    reiwa_year = (
        issue_date.year - 2018
    )

    japanese_date = (
        f"令和{reiwa_year}年"
        f"{issue_date.month}月"
        f"{issue_date.day}日"
    )

    pdf.drawRightString(
        width - 25 * mm,
        height - 32 * mm,
        japanese_date,
    )

    # ============================================================
    # 左側：宛先
    # ============================================================

    address_x = 28 * mm
    address_y = height - 48 * mm

    pdf.setFont(
        font_min,
        11,
    )

    pdf.drawString(
        address_x,
        address_y,
        school.name,
    )

    pdf.drawString(
        address_x + em,
        address_y - 7 * mm,
        "学校長　様",
    )

    pdf.drawString(
        address_x + em,
        address_y - 14 * mm,
        "進学生徒　様",
    )

    # ============================================================
    # 右側：差出人
    # ============================================================

    sender_x = (
        width - 82 * mm
    )

    sender_y = (
        height - 72 * mm
    )

    pdf.setFont(
        font_min,
        10.5,
    )

    pdf.drawString(
        sender_x,
        sender_y,
        "学校法人　高千穂学園",
    )

    pdf.drawString(
        sender_x + em,
        sender_y - 7 * mm,
        "小林西高等学校",
    )

    pdf.drawString(
        sender_x + 2 * em,
        sender_y - 14 * mm,
        f"校長　{principal_name}",
    )

    # ============================================================
    # 校長印
    # ============================================================

    seal_path = finders.find(
        "publicity/images/principal_seal.png"
    )

    if seal_path:

        pdf.drawImage(
            seal_path,
            width - 45 * mm,
            sender_y - 20 * mm,
            width=22 * mm,
            height=22 * mm,
            mask="auto",
            preserveAspectRatio=True,
        )

    # ============================================================
    # 件名
    # ============================================================

    title = (
        f"{recruitment_year}"
        "高千穂学園奨学生の募集について（ご依頼）"
    )

    pdf.setFont(
        font_gothic,
        title_font_size,
    )

    pdf.drawCentredString(
        width / 2,
        title_y,
        title,
    )

    # ============================================================
    # 本文
    # ============================================================

    body_style = ParagraphStyle(
        name=(
            "JapaneseBodyCompact"
            if compact_mode
            else "JapaneseBodyStandard"
        ),
        fontName=font_min,
        fontSize=body_font_size,
        leading=body_leading,
        alignment=TA_LEFT,
        firstLineIndent=(
            body_first_indent
        ),
        spaceAfter=0,
    )

    body_width = (
        width - 46 * mm
    )

    body_x = 23 * mm

    current_y = (
        title_y
        - body_start_gap
    )

    paragraphs = [
        (
            f"謹啓　{seasonal_greeting}、"
            "貴校におかれましてはますますご清栄のことと"
            "お喜び申し上げます。"
            "また、平素より本校の教育活動につきましては、"
            "格別のご厚情を賜り、心より感謝申し上げます。"
        ),
        (
            f"さて、本校では{recruitment_year}"
            "高千穂学園奨学生（文化・スポーツ）を"
            "募集しております。"
        ),
        (
            "そこでこの度、貴校の生徒様を奨学生として"
            "募集いたしたく、ご依頼申し上げる次第です。"
            "つきましては、下記の生徒様ならびに"
            "保護者の皆様へ本件をご案内いただき、"
            "面談の機会を設けていただけますと幸いです。"
        ),
        (
            "誠に恐縮に存じますが、"
            "よろしくお取り計らいくださいますよう"
            "お願い申し上げます。"
        ),
        (
            "なお、御返信に際し校長不在の時は、"
            "教頭が対応いたします。"
            "（TEL 0984-22-5155）"
        ),
    ]

    for text in paragraphs:

        para = Paragraph(
            text,
            body_style,
        )

        _, para_height = (
            para.wrap(
                body_width,
                100 * mm,
            )
        )

        para.drawOn(
            pdf,
            body_x,
            current_y - para_height,
        )

        current_y -= (
            para_height
            + paragraph_gap
        )

    # ============================================================
    # 「記」
    # ============================================================

    current_y -= (
        before_ki_gap
    )

    pdf.setFont(
        font_gothic,
        (
            10
            if compact_mode
            else 11
        ),
    )

    pdf.drawCentredString(
        width / 2,
        current_y,
        "記",
    )

    current_y -= (
        after_ki_gap
    )

    # ============================================================
    # 生徒一覧
    # ============================================================

    table_data = [
        [
            "No.",
            "部活動名",
            "生徒名",
        ]
    ]

    for index, student in enumerate(
        student_list,
        start=1,
    ):

        club_name = ""

        if student.club:
            club_name = (
                student.club.name
            )

        table_data.append(
            [
                str(index),
                club_name,
                f"{student.name} さん",
            ]
        )

    # ============================================================
    # 表
    # ============================================================

    student_table = Table(
        table_data,
        colWidths=[
            18 * mm,
            55 * mm,
            70 * mm,
        ],
        repeatRows=1,
    )

    student_table.setStyle(
        TableStyle(
            [
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, -1),
                    font_min,
                ),
                (
                    "FONTSIZE",
                    (0, 0),
                    (-1, -1),
                    table_font_size,
                ),
                (
                    "LEADING",
                    (0, 0),
                    (-1, -1),
                    table_leading,
                ),
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    font_gothic,
                ),
                (
                    "ALIGN",
                    (0, 0),
                    (-1, 0),
                    "CENTER",
                ),
                (
                    "ALIGN",
                    (0, 1),
                    (0, -1),
                    "CENTER",
                ),
                (
                    "ALIGN",
                    (1, 1),
                    (-1, -1),
                    "LEFT",
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE",
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey,
                ),
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.whitesmoke,
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    table_top_padding,
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    table_bottom_padding,
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    table_left_padding,
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    table_right_padding,
                ),
            ]
        )
    )

    # ============================================================
    # 表サイズ
    # ============================================================

    table_width, table_height = (
        student_table.wrap(
            width - 50 * mm,
            140 * mm,
        )
    )

    # ============================================================
    # 表中央配置
    # ============================================================

    table_x = (
        width - table_width
    ) / 2

    # ============================================================
    # 表の縦位置
    # ============================================================

    table_y = (
        current_y
        - table_height
    )
    
    # ============================================================
    # 安全チェック
    #
    # 20名までの想定だが、
    # 万一ページ下端を超える場合は例外にする。
    # ============================================================

    minimum_bottom = (
        15 * mm
    )

    if table_y < minimum_bottom:

        raise ValueError(
            (
                f"{school.name} は "
                f"{student_count}名のため、"
                "現在の1ページレイアウトに"
                "収まりません。"
            )
        )

    student_table.drawOn(
        pdf,
        table_x,
        table_y,
    )

    # ============================================================
    # 「以上」
    # ============================================================

    current_y = (
        table_y
        - final_gap
    )

    pdf.setFont(
        font_min,
        final_font_size,
    )

    pdf.drawRightString(
        width - 28 * mm,
        current_y,
        "以上",
    )

@club_advisor_required
@transaction.atomic
def scholarship_request_document_pdf(request):

    """
    1校分の奨学生募集依頼文書を
    正式発行しPDFを生成する。
    """

    if request.method != "POST":

        return HttpResponse(
            "不正なアクセスです。",
            status=405,
        )

    # ============================================================
    # フォーム
    # ============================================================

    form = ScholarshipRequestDocumentForm(
        request.POST
    )

    if not form.is_valid():

        return HttpResponse(
            "文書設定に入力エラーがあります。"
            f"<br>{form.errors}",
            status=400,
        )

    # ============================================================
    # 文書基本情報
    # ============================================================

    school = (
        form.cleaned_data["school"]
    )

    issue_date = (
        form.cleaned_data["issue_date"]
    )

    seasonal_greeting = (
        form.cleaned_data[
            "seasonal_greeting"
        ]
    )

    principal_name = (
        form.cleaned_data[
            "principal_name"
        ]
    )

    # ============================================================
    # 募集年度
    #
    # 画面から入力された値を使用
    # 未入力時は発行日時点の年度 + 1
    # ============================================================

    recruitment_year = (
        request.POST
        .get(
            "recruitment_year",
            "",
        )
        .strip()
    )

    if not recruitment_year:

        current_fiscal_year = (
            get_fiscal_year_from_date(
                issue_date
            )
        )

        next_fiscal_year = (
            current_fiscal_year + 1
        )

        next_reiwa_year = (
            next_fiscal_year - 2018
        )

        recruitment_year = (
            f"令和{next_reiwa_year}年度"
        )

    # ============================================================
    # 生徒
    # ============================================================

    students = (
        ProspectiveStudent.objects
        .filter(
            junior_high_school=school,
            is_active=True,
        )
        .select_related(
            "club",
            "junior_high_school",
        )
        .order_by(
            "club__name",
            "name_kana",
            "name",
        )
    )

    if not students.exists():

        return HttpResponse(
            "対象生徒が登録されていません。",
            status=400,
        )

    # ============================================================
    # 文書管理年度
    # ============================================================

    fiscal_year = (
        get_fiscal_year_from_date(
            issue_date
        )
    )

    # ============================================================
    # 正式採番
    # ============================================================

    sequence, _ = (
        DocumentNumberSequence.objects
        .select_for_update()
        .get_or_create(
            fiscal_year=fiscal_year,
            defaults={
                "last_number": 0,
            },
        )
    )

    sequence.last_number += 1

    sequence.save(
        update_fields=[
            "last_number",
            "updated_at",
        ]
    )

    document_number = (
        f"小林西発第"
        f"{sequence.last_number:03d}"
        f"号"
    )

    # ============================================================
    # 発行履歴保存
    # ============================================================

    document = (
        ScholarshipRequestDocument.objects
        .create(
            school=school,

            # 文書管理年度
            fiscal_year=fiscal_year,

            # 募集年度
            recruitment_year=(
                recruitment_year
            ),

            document_number=(
                document_number
            ),

            issue_date=(
                issue_date
            ),

            seasonal_greeting=(
                seasonal_greeting
            ),

            principal_name=(
                principal_name
            ),

            created_by=(
                request.teacher
            ),

            status="issued",
        )
    )

    document.students.set(
        students
    )

    # ============================================================
    # PDF作成
    # ============================================================

    buffer = BytesIO()

    pdf = canvas.Canvas(
        buffer,
        pagesize=A4,
    )

    draw_scholarship_request_page(
        pdf=pdf,

        school=school,

        students=students,

        document_number=(
            document_number
        ),

        issue_date=(
            issue_date
        ),

        # 文書管理年度
        fiscal_year=(
            fiscal_year
        ),

        # 募集年度
        recruitment_year=(
            recruitment_year
        ),

        seasonal_greeting=(
            seasonal_greeting
        ),

        principal_name=(
            principal_name
        ),
    )

    pdf.showPage()

    pdf.save()

    buffer.seek(0)

    # ============================================================
    # PDF返却
    # ============================================================

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/pdf",
    )

    filename = (
        f"{document_number}_"
        f"{school.name}.pdf"
    )

    response[
        "Content-Disposition"
    ] = (
        f'inline; filename="{quote(filename)}"'
    )

    return response

@system_admin_required
@transaction.atomic
def scholarship_document_batch_issue(request):
    """
    選択された生徒を中学校ごとにまとめ、
    1つの共通文書番号を使用して
    複数ページPDFとして正式発行する。

    ・文書番号は今回の発行処理で1つ
    ・複数校でも同じ文書番号を使用
    ・文書番号は事務と連携して手入力
    ・募集年度は画面入力値を使用
    ・募集年度未入力時は翌年度を自動設定
    ・1校あたり最大10名
    ・すべて検証後にDB保存
    """

    if request.method != "POST":

        return redirect(
            "publicity:scholarship_document_student_select"
        )

    # ============================================================
    # POSTデータ
    # ============================================================

    student_ids = request.POST.getlist(
        "student_ids"
    )

    issue_date_text = request.POST.get(
        "issue_date",
        "",
    ).strip()

    seasonal_greeting = request.POST.get(
        "seasonal_greeting",
        "",
    ).strip()

    principal_name = request.POST.get(
        "principal_name",
        "",
    ).strip()

    document_number = request.POST.get(
        "document_number",
        "",
    ).strip()

    # ============================================================
    # 募集年度
    #
    # 画面から入力された
    # 「令和9年度」等を受け取る
    # ============================================================

    recruitment_year = request.POST.get(
        "recruitment_year",
        "",
    ).strip()

    # ------------------------------------------------------------
    # 募集年度未入力時の保険
    # 現在年度 + 1
    # ------------------------------------------------------------

    if not recruitment_year:

        today = timezone.localdate()

        current_fiscal_year = (
            get_fiscal_year_from_date(
                today
            )
        )

        next_fiscal_year = (
            current_fiscal_year + 1
        )

        next_reiwa_year = (
            next_fiscal_year - 2018
        )

        recruitment_year = (
            f"令和{next_reiwa_year}年度"
        )

    # ============================================================
    # 必須項目チェック
    # ============================================================

    if not student_ids:

        messages.error(
            request,
            "対象生徒が選択されていません。",
        )

        return redirect(
            "publicity:scholarship_document_student_select"
        )

    if not document_number:

        return HttpResponse(
            "文書番号が入力されていません。",
            status=400,
        )

    if not issue_date_text:

        return HttpResponse(
            "発行日が指定されていません。",
            status=400,
        )

    if not seasonal_greeting:

        return HttpResponse(
            "時候の挨拶が指定されていません。",
            status=400,
        )

    if not principal_name:

        return HttpResponse(
            "校長名が指定されていません。",
            status=400,
        )

    # ============================================================
    # 発行日
    # ============================================================

    try:

        issue_date = datetime.strptime(
            issue_date_text,
            "%Y-%m-%d",
        ).date()

    except ValueError:

        return HttpResponse(
            "発行日の形式が正しくありません。",
            status=400,
        )

    # ============================================================
    # 対象生徒取得
    # ============================================================

    students = list(
        ProspectiveStudent.objects
        .filter(
            id__in=student_ids,
            is_active=True,
        )
        .select_related(
            "junior_high_school",
            "club",
            "assigned_teacher",
        )
        .order_by(
            "junior_high_school__prefecture",
            "junior_high_school__city",
            "junior_high_school__name",
            "club__name",
            "name",
        )
    )

    if not students:

        return HttpResponse(
            "有効な対象生徒が見つかりません。",
            status=400,
        )

    # ============================================================
    # 中学校別にまとめる
    # ============================================================

    school_groups = {}

    for student in students:

        school = (
            student.junior_high_school
        )

        if school.id not in school_groups:

            school_groups[
                school.id
            ] = {
                "school": school,
                "students": [],
            }

        school_groups[
            school.id
        ]["students"].append(
            student
        )

    groups = list(
        school_groups.values()
    )

    # ============================================================
    # 1校10名まで
    # ============================================================

    for group in groups:

        school = group["school"]

        school_students = (
            group["students"]
        )

        if len(school_students) > 16:

            return HttpResponse(
                f"{school.name}の対象生徒が"
                "16名を超えています。"
                "1文書につき16名以内で"
                "選択してください。",
                status=400,
            )

    # ============================================================
    # 文書管理年度
    #
    # これは「募集年度」ではない。
    # 文書番号・履歴管理用。
    # ============================================================

    fiscal_year = (
        get_fiscal_year_from_date(
            issue_date
        )
    )

    # ============================================================
    # 文書番号の既存チェック
    # ============================================================

    school_ids = [
        group["school"].id
        for group in groups
    ]

    duplicate_documents = (
        ScholarshipRequestDocument.objects
        .filter(
            document_number=document_number,
            issue_date=issue_date,
            school_id__in=school_ids,
        )
    )

    if duplicate_documents.exists():

        duplicate_school_names = (
            duplicate_documents
            .values_list(
                "school__name",
                flat=True,
            )
        )

        duplicate_text = "、".join(
            duplicate_school_names
        )

        return HttpResponse(
            "同じ文書がすでに発行されています。"
            f"<br>文書番号：{document_number}"
            f"<br>発行日：{issue_date}"
            f"<br>対象校：{duplicate_text}",
            status=400,
        )

    # ============================================================
    # PDF準備
    # ============================================================

    buffer = BytesIO()

    pdf = canvas.Canvas(
        buffer,
        pagesize=A4,
    )

    # ============================================================
    # 正式発行
    # ============================================================

    for group in groups:

        school = group["school"]

        school_students = (
            group["students"]
        )

        # --------------------------------------------------------
        # 発行履歴保存
        # --------------------------------------------------------

        document = (
            ScholarshipRequestDocument.objects
            .create(
                school=school,

                # 文書管理年度
                fiscal_year=fiscal_year,

                # 募集年度
                recruitment_year=recruitment_year,

                document_number=document_number,

                issue_date=issue_date,

                seasonal_greeting=(
                    seasonal_greeting
                ),

                principal_name=(
                    principal_name
                ),

                created_by=request.teacher,

                status="issued",
            )
        )

        document.students.set(
            school_students
        )

        # --------------------------------------------------------
        # PDF描画
        # --------------------------------------------------------

        draw_scholarship_request_page(
            pdf=pdf,

            school=school,

            students=school_students,

            document_number=(
                document_number
            ),

            issue_date=issue_date,

            # 文書管理年度
            fiscal_year=fiscal_year,

            # ★募集年度
            recruitment_year=(
                recruitment_year
            ),

            seasonal_greeting=(
                seasonal_greeting
            ),

            principal_name=(
                principal_name
            ),
        )

        # 1校につき1ページ
        pdf.showPage()

    # ============================================================
    # Sequence更新
    # ============================================================

    match = re.search(
        r"第\s*(\d+)\s*号",
        document_number,
    )

    if match:

        numeric_number = int(
            match.group(1)
        )

        sequence, _ = (
            DocumentNumberSequence.objects
            .select_for_update()
            .get_or_create(
                fiscal_year=fiscal_year,
                defaults={
                    "last_number": 0,
                },
            )
        )

        if (
            numeric_number
            > sequence.last_number
        ):

            sequence.last_number = (
                numeric_number
            )

            sequence.save(
                update_fields=[
                    "last_number",
                    "updated_at",
                ]
            )

    # ============================================================
    # PDF完了
    # ============================================================

    pdf.save()

    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/pdf",
    )

    filename = (
        "高千穂学園奨学生募集依頼_"
        f"{document_number}_"
        f"{issue_date.strftime('%Y%m%d')}_"
        f"{len(groups)}校.pdf"
    )

    response[
        "Content-Disposition"
    ] = (
        f'inline; filename="{quote(filename)}"'
    )

    return response

@system_admin_required
def scholarship_document_history(request):

    documents = (
        ScholarshipRequestDocument.objects
        .select_related(
            "school",
            "created_by",
        )
        .prefetch_related(
            "students",
        )
        .order_by(
            "-issue_date",
            "-created_at",
        )
    )

    keyword = request.GET.get(
        "q",
        "",
    ).strip()

    status = request.GET.get(
        "status",
        "",
    ).strip()

    if keyword:
        documents = documents.filter(
            Q(document_number__icontains=keyword)
            | Q(recruitment_year__icontains=keyword)
            | Q(school__name__icontains=keyword)
            | Q(students__name__icontains=keyword)
        ).distinct()

    if status:
        documents = documents.filter(
            status=status
        )

    return render(
        request,
        "publicity/scholarship_document_history.html",
        {
            "documents": documents,
            "keyword": keyword,
            "selected_status": status,
        },
    )

@system_admin_required
def scholarship_document_reprint(
    request,
    document_id,
):

    # ============================================================
    # 対象文書取得
    # ============================================================

    document = get_object_or_404(
        ScholarshipRequestDocument,
        pk=document_id,
    )

    # ============================================================
    # 対象生徒取得
    # ============================================================

    students = (
        document.students
        .select_related(
            "club",
            "junior_high_school",
        )
        .order_by(
            "club__name",
            "name_kana",
            "name",
        )
    )

    # ============================================================
    # 募集年度
    #
    # 新しい履歴：
    # DBに保存された recruitment_year を使用
    #
    # 古い履歴：
    # recruitment_year が空欄なら、
    # 文書管理年度 + 1 から自動生成
    # ============================================================

    recruitment_year = (
        document.recruitment_year
        or ""
    ).strip()

    if not recruitment_year:

        recruitment_fiscal_year = (
            document.fiscal_year + 1
        )

        recruitment_reiwa_year = (
            recruitment_fiscal_year - 2018
        )

        recruitment_year = (
            f"令和{recruitment_reiwa_year}年度"
        )

    # ============================================================
    # PDF準備
    # ============================================================

    buffer = BytesIO()

    pdf = canvas.Canvas(
        buffer,
        pagesize=A4,
    )

    # ============================================================
    # PDF描画
    # ============================================================

    draw_scholarship_request_page(
        pdf=pdf,

        school=document.school,

        students=students,

        document_number=(
            document.document_number
        ),

        issue_date=(
            document.issue_date
        ),

        # 文書管理年度
        fiscal_year=(
            document.fiscal_year
        ),

        # 発行時に保存した募集年度
        recruitment_year=(
            recruitment_year
        ),

        seasonal_greeting=(
            document.seasonal_greeting
        ),

        principal_name=(
            document.principal_name
        ),
    )

    pdf.showPage()

    pdf.save()

    buffer.seek(0)

    # ============================================================
    # PDF返却
    # ============================================================

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/pdf",
    )

    filename = (
        f"{document.document_number}_"
        f"{document.school.name}_再印刷.pdf"
    )

    response[
        "Content-Disposition"
    ] = (
        f'inline; filename="{quote(filename)}"'
    )

    return response

@system_admin_required
@transaction.atomic
def scholarship_document_cancel(
    request,
    document_id,
):

    document = get_object_or_404(
        ScholarshipRequestDocument,
        pk=document_id,
    )

    if request.method != "POST":
        return HttpResponse(
            "不正なアクセスです。",
            status=405,
        )

    # すでに取消済みなら何もしない
    if document.status == "cancelled":

        messages.info(
            request,
            "この文書はすでに取消済みです。",
        )

        return redirect(
            "publicity:scholarship_document_history"
        )

    document.status = "cancelled"

    document.save(
        update_fields=[
            "status",
        ]
    )

    messages.success(
        request,
        (
            f"{document.document_number} "
            f"{document.school.name} の文書を"
            "取消しました。"
        ),
    )

    return redirect(
        "publicity:scholarship_document_history"
    )

@system_admin_required
@transaction.atomic
def scholarship_document_correct(
    request,
    document_id,
):

    # ============================================================
    # 元文書取得
    # ============================================================

    original = get_object_or_404(
        ScholarshipRequestDocument,
        pk=document_id,
    )

    # ============================================================
    # 取消済みは訂正不可
    # ============================================================

    if original.status == "cancelled":

        messages.error(
            request,
            "取消済みの文書は訂正できません。",
        )

        return redirect(
            "publicity:scholarship_document_history"
        )

    # ============================================================
    # 旧版は再訂正しない
    # ============================================================

    if original.status == "corrected":

        messages.error(
            request,
            "この文書はすでに訂正済みです。"
            "最新の訂正版を訂正してください。",
        )

        return redirect(
            "publicity:scholarship_document_history"
        )

    # ============================================================
    # 元文書の募集年度
    #
    # 新しい履歴：
    # original.recruitment_year を使用
    #
    # 古い履歴：
    # 空欄なら fiscal_year + 1 から補完
    # ============================================================

    recruitment_year = (
        original.recruitment_year
        or ""
    ).strip()

    if not recruitment_year:

        recruitment_fiscal_year = (
            original.fiscal_year + 1
        )

        recruitment_reiwa_year = (
            recruitment_fiscal_year - 2018
        )

        recruitment_year = (
            f"令和{recruitment_reiwa_year}年度"
        )

    # ============================================================
    # 元文書の対象生徒
    # ============================================================

    original_students = list(
        original.students
        .select_related(
            "club",
            "junior_high_school",
        )
        .order_by(
            "club__name",
            "name_kana",
            "name",
        )
    )

    # ============================================================
    # 同中学校の現在有効な生徒
    # ============================================================

    available_students = list(
        ProspectiveStudent.objects
        .filter(
            junior_high_school=original.school,
            is_active=True,
        )
        .select_related(
            "club",
            "junior_high_school",
        )
        .order_by(
            "club__name",
            "name_kana",
            "name",
        )
    )

    selected_student_ids = {
        student.id
        for student in original_students
    }

    # ============================================================
    # POST：訂正版作成
    # ============================================================

    if request.method == "POST":

        document_number = request.POST.get(
            "document_number",
            "",
        ).strip()

        issue_date_text = request.POST.get(
            "issue_date",
            "",
        ).strip()

        seasonal_greeting = request.POST.get(
            "seasonal_greeting",
            "",
        ).strip()

        principal_name = request.POST.get(
            "principal_name",
            "",
        ).strip()

        correction_reason = request.POST.get(
            "correction_reason",
            "",
        ).strip()

        student_ids = request.POST.getlist(
            "student_ids"
        )

        # ========================================================
        # 必須チェック
        # ========================================================

        if not document_number:

            messages.error(
                request,
                "文書番号を入力してください。",
            )

            return redirect(
                "publicity:scholarship_document_correct",
                document_id=original.id,
            )

        if not issue_date_text:

            messages.error(
                request,
                "発行日を入力してください。",
            )

            return redirect(
                "publicity:scholarship_document_correct",
                document_id=original.id,
            )

        if not seasonal_greeting:

            messages.error(
                request,
                "時候の挨拶を入力してください。",
            )

            return redirect(
                "publicity:scholarship_document_correct",
                document_id=original.id,
            )

        if not principal_name:

            messages.error(
                request,
                "校長名を入力してください。",
            )

            return redirect(
                "publicity:scholarship_document_correct",
                document_id=original.id,
            )

        if not correction_reason:

            messages.error(
                request,
                "訂正理由を入力してください。",
            )

            return redirect(
                "publicity:scholarship_document_correct",
                document_id=original.id,
            )

        # ========================================================
        # 発行日
        # ========================================================

        try:

            issue_date = datetime.strptime(
                issue_date_text,
                "%Y-%m-%d",
            ).date()

        except ValueError:

            messages.error(
                request,
                "発行日の形式が正しくありません。",
            )

            return redirect(
                "publicity:scholarship_document_correct",
                document_id=original.id,
            )

        # ========================================================
        # 対象生徒
        #
        # 画面で選択があればそれを採用
        # 何も取れなければ元文書を引き継ぐ
        # ========================================================

        if student_ids:

            correction_students = list(
                ProspectiveStudent.objects
                .filter(
                    id__in=student_ids,
                    junior_high_school=original.school,
                    is_active=True,
                )
                .select_related(
                    "club",
                    "junior_high_school",
                )
                .order_by(
                    "club__name",
                    "name_kana",
                    "name",
                )
            )

        else:

            correction_students = (
                original_students
            )

        if not correction_students:

            messages.error(
                request,
                "訂正版に掲載する生徒がいません。",
            )

            return redirect(
                "publicity:scholarship_document_correct",
                document_id=original.id,
            )

        if len(correction_students) > 10:

            messages.error(
                request,
                "対象生徒は10名以内にしてください。",
            )

            return redirect(
                "publicity:scholarship_document_correct",
                document_id=original.id,
            )

        # ========================================================
        # 文書管理年度
        #
        # 訂正版の発行日から再計算
        # ========================================================

        fiscal_year = (
            get_fiscal_year_from_date(
                issue_date
            )
        )

        # ========================================================
        # 新しいDBレコードとして訂正版作成
        # ========================================================

        corrected_document = (
            ScholarshipRequestDocument.objects
            .create(
                school=original.school,

                # 文書管理年度
                fiscal_year=fiscal_year,

                # 元文書の募集年度を引継ぎ
                recruitment_year=(
                    recruitment_year
                ),

                document_number=(
                    document_number
                ),

                issue_date=(
                    issue_date
                ),

                seasonal_greeting=(
                    seasonal_greeting
                ),

                principal_name=(
                    principal_name
                ),

                created_by=(
                    request.teacher
                ),

                status="issued",

                corrected_from=(
                    original
                ),

                correction_reason=(
                    correction_reason
                ),
            )
        )

        corrected_document.students.set(
            correction_students
        )

        # ========================================================
        # 元文書を訂正済みに変更
        # ========================================================

        original.status = "corrected"

        original.save(
            update_fields=[
                "status",
            ]
        )

        messages.success(
            request,
            (
                f"{original.document_number} "
                f"{original.school.name} の訂正版を"
                f"作成しました。"
                f"新しい内部IDは "
                f"{corrected_document.id} です。"
            ),
        )

        return redirect(
            "publicity:scholarship_document_history"
        )

    # ============================================================
    # GET：訂正画面表示
    # ============================================================

    return render(
        request,
        "publicity/scholarship_document_correct.html",
        {
            "original": original,

            "selected_students": (
                original_students
            ),

            "selected_student_ids": (
                selected_student_ids
            ),

            "available_students": (
                available_students
            ),

            # 訂正画面でも確認できるように渡す
            "recruitment_year": (
                recruitment_year
            ),
        },
    )

# ============================================================
# 奨学生・面談管理一覧
# ============================================================

@club_advisor_required
def scholarship_management_list(request):

    """
    部顧問用

    奨学生候補・面談管理一覧

    原則として、
    ログイン中の教員が担当している募集対象生徒のみ表示する。

    また、
    担当部活動ごとの奨学生枠・寮費待遇枠について、
    現在人員 / 上限人数を表示する。
    """

    # --------------------------------------------------------
    # 検索条件
    # --------------------------------------------------------

    keyword = (
        request.GET
        .get("q", "")
        .strip()
    )

    club_id = (
        request.GET
        .get("club", "")
        .strip()
    )

    status = (
        request.GET
        .get("status", "")
        .strip()
    )

    # --------------------------------------------------------
    # 対象年度
    #
    # 現在は「来年度募集」を基準とする
    # 例：
    # 2026年中の募集 → 2027年度
    # --------------------------------------------------------

    today = timezone.localdate()

    fiscal_year = today.year + 1

    # --------------------------------------------------------
    # ログイン教員が担当している全生徒
    #
    # 枠使用数の集計はこちらを使用する。
    # 検索条件によって枠使用数が変わらないようにするため。
    # --------------------------------------------------------

    all_students = (
        ProspectiveStudent.objects
        .filter(
            is_active=True,
            assigned_teacher=request.teacher,
        )
        .select_related(
            "junior_high_school",
            "club",
            "assigned_teacher",

            "scholarship_assignment",

            "scholarship_assignment__interview_rank",
            "scholarship_assignment__current_rank",
            "scholarship_assignment__final_rank",

            "scholarship_assignment__interview_dormitory_benefit",
            "scholarship_assignment__current_dormitory_benefit",
            "scholarship_assignment__final_dormitory_benefit",
        )
    )

    # --------------------------------------------------------
    # 一覧表示用QuerySet
    # --------------------------------------------------------

    students = (
        all_students
        .order_by(
            "club__name",
            "junior_high_school__name",
            "name_kana",
            "name",
        )
    )

    # --------------------------------------------------------
    # キーワード検索
    # --------------------------------------------------------

    if keyword:

        students = students.filter(
            Q(
                name__icontains=keyword
            )
            | Q(
                name_kana__icontains=keyword
            )
            | Q(
                junior_high_school__name__icontains=keyword
            )
            | Q(
                club__name__icontains=keyword
            )
        )

    # --------------------------------------------------------
    # 部活動絞り込み
    # --------------------------------------------------------

    if club_id.isdigit():

        students = students.filter(
            club_id=club_id
        )

    # --------------------------------------------------------
    # 進行状況絞り込み
    # --------------------------------------------------------

    if status:

        if status == "not_started":

            students = students.filter(
                Q(
                    scholarship_assignment__isnull=True
                )
                | Q(
                    scholarship_assignment__status="not_started"
                )
            )

        else:

            students = students.filter(
                scholarship_assignment__status=status
            )

    # --------------------------------------------------------
    # 部活動選択肢
    # --------------------------------------------------------

    clubs = (
        Club.objects
        .filter(
            prospective_students__assigned_teacher=request.teacher,
            prospective_students__is_active=True,
        )
        .distinct()
        .order_by(
            "name"
        )
    )

    # --------------------------------------------------------
    # 一覧側 件数集計
    # --------------------------------------------------------

    total_count = students.count()

    not_started_count = (
        students.filter(
            Q(
                scholarship_assignment__isnull=True
            )
            | Q(
                scholarship_assignment__status="not_started"
            )
        )
        .count()
    )

    interview_count = (
        students.filter(
            scholarship_assignment__status__in=[
                "interview_scheduled",
                "interviewed",
                "temporary_accepted",
            ]
        )
        .count()
    )

    adjusting_count = (
        students.filter(
            scholarship_assignment__status="adjusting"
        )
        .count()
    )

    finalized_count = (
        students.filter(
            scholarship_assignment__status__in=[
                "conference_confirmed",
                "finalized",
                "accepted",
            ]
        )
        .count()
    )

    # ========================================================
    # 奨学生ランク枠集計
    # ========================================================

    categories = (
        ScholarshipCategory.objects
        .filter(
            is_active=True
        )
        .annotate(
            display_order=Case(
                When(
                    name="SS",
                    then=Value(1),
                ),
                When(
                    name="S",
                    then=Value(2),
                ),
                When(
                    name="A",
                    then=Value(3),
                ),
                default=Value(99),
                output_field=IntegerField(),
            )
        )
        .order_by(
            "display_order",
            "name",
        )
    )

    # --------------------------------------------------------
    # 寮費枠対象
    #
    # 実際のマスタ名称に合わせること
    # --------------------------------------------------------

    dormitory_quota_benefit = (
        DormitoryBenefitCategory.objects
        .filter(
            is_active=True,
            name="寮施設管理費免除",
        )
        .first()
    )

    # --------------------------------------------------------
    # 部活動別 枠利用状況
    # --------------------------------------------------------

    quota_summary_rows = []

    for club in clubs:

        # ----------------------------------------------------
        # この部活動の担当生徒全体
        # ----------------------------------------------------

        club_students = all_students.filter(
            club=club
        )

        # ----------------------------------------------------
        # ランク別
        # ----------------------------------------------------

        rank_summaries = []

        for category in categories:

            current_count = (
                club_students.filter(
                    scholarship_assignment__current_rank=category
                )
                .count()
            )

            quota_record = (
                ScholarshipQuota.objects
                .filter(
                    fiscal_year=fiscal_year,
                    club=club,
                    category=category,
                )
                .first()
            )

            quota_count = (
                quota_record.quota
                if quota_record
                else 0
            )

            remaining = (
                quota_count - current_count
            )

            rank_summaries.append(
                {
                    "category": category,
                    "current_count": current_count,
                    "quota_count": quota_count,
                    "remaining": remaining,
                    "is_full": (
                        quota_count > 0
                        and current_count == quota_count
                    ),
                    "is_over": (
                        quota_count >= 0
                        and current_count > quota_count
                    ),
                }
            )

        # ----------------------------------------------------
        # 寮施設管理費免除
        # ----------------------------------------------------

        dormitory_current_count = 0
        dormitory_quota_count = 0

        if dormitory_quota_benefit is not None:

            dormitory_current_count = (
                club_students.filter(
                    scholarship_assignment__current_dormitory_benefit=(
                        dormitory_quota_benefit
                    )
                )
                .count()
            )

            dormitory_quota_record = (
                DormitoryBenefitQuota.objects
                .filter(
                    fiscal_year=fiscal_year,
                    club=club,
                    benefit=dormitory_quota_benefit,
                )
                .first()
            )

            if dormitory_quota_record:

                dormitory_quota_count = (
                    dormitory_quota_record.quota
                )

        dormitory_remaining = (
            dormitory_quota_count
            - dormitory_current_count
        )

        # ----------------------------------------------------
        # 1部活動分
        # ----------------------------------------------------

        quota_summary_rows.append(
            {
                "club": club,
                "rank_summaries": rank_summaries,

                "dormitory_current_count":
                    dormitory_current_count,

                "dormitory_quota_count":
                    dormitory_quota_count,

                "dormitory_remaining":
                    dormitory_remaining,

                "dormitory_is_full": (
                    dormitory_quota_count > 0
                    and dormitory_current_count
                    == dormitory_quota_count
                ),

                "dormitory_is_over": (
                    dormitory_quota_count >= 0
                    and dormitory_current_count
                    > dormitory_quota_count
                ),
            }
        )

    # --------------------------------------------------------
    # ステータス選択肢
    # --------------------------------------------------------

    status_choices = [
        (
            "",
            "すべて",
        ),
        (
            "not_started",
            "未対応",
        ),
        (
            "rank_set",
            "ランク設定済",
        ),
        (
            "interview_scheduled",
            "面談予定",
        ),
        (
            "interviewed",
            "面談済",
        ),
        (
            "temporary_accepted",
            "仮承諾済",
        ),
        (
            "adjusting",
            "ランク調整中",
        ),
        (
            "conference_confirmed",
            "連絡会確認済",
        ),
        (
            "finalized",
            "最終確定",
        ),
        (
            "accepted",
            "正式承諾",
        ),
        (
            "declined",
            "辞退",
        ),
    ]

    # --------------------------------------------------------
    # Template
    # --------------------------------------------------------

    context = {

        "students": students,

        "clubs": clubs,

        "keyword": keyword,

        "selected_club": club_id,

        "selected_status": status,

        "status_choices": status_choices,

        # 一覧件数
        "total_count": total_count,

        "not_started_count":
            not_started_count,

        "interview_count":
            interview_count,

        "adjusting_count":
            adjusting_count,

        "finalized_count":
            finalized_count,

        # 枠管理
        "fiscal_year":
            fiscal_year,

        "quota_summary_rows":
            quota_summary_rows,

        "dormitory_quota_benefit":
            dormitory_quota_benefit,
    }

    return render(
        request,
        "publicity/scholarship_management_list.html",
        context,
    )

# ============================================================
# 奨学生・面談 個人管理
# ============================================================

@club_advisor_required
def scholarship_assignment_detail(request, student_id):

    teacher = request.teacher

    # --------------------------------------------------------
    # 対象生徒取得
    #
    # 部顧問
    #   → 自分の担当生徒のみ
    #
    # 広報管理者・システム管理者
    #   → 全生徒
    # --------------------------------------------------------

    student_queryset = (
        ProspectiveStudent.objects
        .select_related(
            "junior_high_school",
            "club",
            "assigned_teacher",
            "registered_by",
        )
        .filter(
            is_active=True,
        )
    )

    if teacher.role not in {
        "publicity_admin",
        "system_admin",
    }:

        student_queryset = student_queryset.filter(
            Q(assigned_teacher=teacher)
            | Q(registered_by=teacher)
        )

    student = get_object_or_404(
        student_queryset,
        id=student_id,
    )


    # --------------------------------------------------------
    # 奨学生管理レコード取得
    #
    # 初回アクセス時にはまだ存在しないため、
    # ここで自動作成する
    # --------------------------------------------------------

    assignment, created = (
        ScholarshipAssignment.objects
        .get_or_create(
            student=student,
        )
    )


    # --------------------------------------------------------
    # 使用中の奨学金ランク
    # --------------------------------------------------------

    scholarship_categories = (
        ScholarshipCategory.objects
        .filter(
            is_active=True
        )
        .order_by(
            "name"
        )
    )


    # --------------------------------------------------------
    # 使用中の寮費区分
    # --------------------------------------------------------

    dormitory_categories = (
        DormitoryBenefitCategory.objects
        .filter(
            is_active=True
        )
        .order_by(
            "name"
        )
    )


    # --------------------------------------------------------
    # 面談履歴
    # --------------------------------------------------------

    interviews = (
        assignment.interviews
        .select_related(
            "interviewer",
            "presented_rank",
            "presented_dormitory_benefit",
        )
        .all()
    )


    # --------------------------------------------------------
    # ランク変更履歴
    # --------------------------------------------------------

    rank_histories = (
        assignment.rank_histories
        .select_related(
            "previous_rank",
            "new_rank",
            "changed_by",
        )
        .all()
    )


    # --------------------------------------------------------
    # 寮費変更履歴
    # --------------------------------------------------------

    dormitory_histories = (
        assignment.dormitory_benefit_histories
        .select_related(
            "previous_benefit",
            "new_benefit",
            "changed_by",
        )
        .all()
    )


    # --------------------------------------------------------
    # 面談前条件の登録
    # --------------------------------------------------------

    if request.method == "POST":

        action = request.POST.get(
            "action"
        )

        if action == "set_initial_conditions":

            rank_id = request.POST.get(
                "interview_rank"
            )

            dormitory_id = request.POST.get(
                "interview_dormitory_benefit"
            )

            # ----------------------------------------------------
            # 変更前の値を保持
            # ----------------------------------------------------

            old_rank = assignment.current_rank

            old_dormitory_benefit = (
                assignment.current_dormitory_benefit
            )

            # ----------------------------------------------------
            # 新しいランク取得
            # ----------------------------------------------------

            rank = None

            if rank_id:

                rank = get_object_or_404(
                    ScholarshipCategory,
                    id=rank_id,
                    is_active=True,
                )

            # ----------------------------------------------------
            # 新しい寮費区分取得
            # ----------------------------------------------------

            dormitory_benefit = None

            if dormitory_id:

                dormitory_benefit = get_object_or_404(
                    DormitoryBenefitCategory,
                    id=dormitory_id,
                    is_active=True,
                )

            # ----------------------------------------------------
            # ランク変更履歴
            # ----------------------------------------------------

            if old_rank != rank:

                # 新しいランクが設定される場合のみ履歴を作る
                if rank is not None:

                    ScholarshipRankHistory.objects.create(
                        assignment=assignment,
                        previous_rank=old_rank,
                        new_rank=rank,
                        reason="面談前条件の設定・変更",
                        changed_by=teacher,
                    )

            # ----------------------------------------------------
            # 寮費区分変更履歴
            # ----------------------------------------------------

            if old_dormitory_benefit != dormitory_benefit:

                # 新しい寮費区分が設定される場合のみ履歴を作る
                if dormitory_benefit is not None:

                    DormitoryBenefitHistory.objects.create(
                        assignment=assignment,
                        previous_benefit=old_dormitory_benefit,
                        new_benefit=dormitory_benefit,
                        reason="面談前条件の設定・変更",
                        changed_by=teacher,
                    )

            # ----------------------------------------------------
            # 面談前条件を保存
            # ----------------------------------------------------

            assignment.interview_rank = rank

            assignment.current_rank = rank

            assignment.interview_dormitory_benefit = (
                dormitory_benefit
            )

            assignment.current_dormitory_benefit = (
                dormitory_benefit
            )

            if rank is not None:

                assignment.status = "rank_set"

            else:

                assignment.status = "not_started"

            assignment.save()

            messages.success(
                request,
                "面談前の奨学生条件を保存しました。",
            )

            return redirect(
                "publicity:scholarship_assignment_detail",
                student_id=student.id,
            )
                # ====================================================
        # 面談後の条件調整
        # ====================================================

        elif action == "adjust_conditions":

            rank_id = (
                request.POST
                .get(
                    "current_rank",
                    "",
                )
                .strip()
            )

            dormitory_id = (
                request.POST
                .get(
                    "current_dormitory_benefit",
                    "",
                )
                .strip()
            )

            rank_reason = (
                request.POST
                .get(
                    "rank_adjustment_reason",
                    "",
                )
                .strip()
            )

            dormitory_reason = (
                request.POST
                .get(
                    "dormitory_adjustment_reason",
                    "",
                )
                .strip()
            )

            # ------------------------------------------------
            # 現在値を保持
            # ------------------------------------------------

            previous_rank = (
                assignment.current_rank
            )

            previous_dormitory = (
                assignment.current_dormitory_benefit
            )

            # ------------------------------------------------
            # 新しいランク
            # ------------------------------------------------

            new_rank = None

            if rank_id:

                new_rank = get_object_or_404(
                    ScholarshipCategory,
                    id=rank_id,
                    is_active=True,
                )

            # ------------------------------------------------
            # 新しい寮費区分
            # ------------------------------------------------

            new_dormitory = None

            if dormitory_id:

                new_dormitory = get_object_or_404(
                    DormitoryBenefitCategory,
                    id=dormitory_id,
                    is_active=True,
                )

            # ------------------------------------------------
            # 何も変更されていない場合
            # ------------------------------------------------

            if (
                previous_rank == new_rank
                and previous_dormitory == new_dormitory
            ):

                messages.info(
                    request,
                    "奨学生条件に変更はありませんでした。",
                )

                return redirect(
                    "publicity:scholarship_assignment_detail",
                    student_id=student.id,
                )

            # ------------------------------------------------
            # ランク変更
            # ------------------------------------------------

            if previous_rank != new_rank:

                if new_rank is None:

                    messages.error(
                        request,
                        "現在ランクを未設定へ戻すことは"
                        "できません。",
                    )

                    return redirect(
                        "publicity:scholarship_assignment_detail",
                        student_id=student.id,
                    )

                ScholarshipRankHistory.objects.create(
                    assignment=assignment,
                    previous_rank=previous_rank,
                    new_rank=new_rank,
                    reason=(
                        rank_reason
                        or "面談後のランク調整"
                    ),
                    changed_by=teacher,
                )

                assignment.current_rank = (
                    new_rank
                )

            # ------------------------------------------------
            # 寮費区分変更
            # ------------------------------------------------

            if (
                previous_dormitory
                != new_dormitory
            ):

                # 現在の履歴モデルでは
                # new_benefit がNULL不可のため、
                # 未設定への戻しは対象外
                if new_dormitory is None:

                    messages.error(
                        request,
                        "寮費区分を未設定へ戻すことは"
                        "現在できません。",
                    )

                    return redirect(
                        "publicity:scholarship_assignment_detail",
                        student_id=student.id,
                    )

                DormitoryBenefitHistory.objects.create(
                    assignment=assignment,
                    previous_benefit=(
                        previous_dormitory
                    ),
                    new_benefit=(
                        new_dormitory
                    ),
                    reason=(
                        dormitory_reason
                        or "面談後の寮費条件調整"
                    ),
                    changed_by=teacher,
                )

                assignment.current_dormitory_benefit = (
                    new_dormitory
                )

            # ------------------------------------------------
            # 進行状況
            # ------------------------------------------------

            assignment.status = "adjusting"

            assignment.save()

            messages.success(
                request,
                "面談後の奨学生条件を更新しました。",
            )

            return redirect(
                "publicity:scholarship_assignment_detail",
                student_id=student.id,
            )


    # --------------------------------------------------------
    # Template
    # --------------------------------------------------------

    context = {

        "teacher": teacher,

        "student": student,

        "assignment": assignment,

        "scholarship_categories": (
            scholarship_categories
        ),

        "dormitory_categories": (
            dormitory_categories
        ),

        "interviews": interviews,

        "rank_histories": (
            rank_histories
        ),

        "dormitory_histories": (
            dormitory_histories
        ),

        "assignment_created": created,
    }


    return render(
        request,
        "publicity/scholarship_assignment_detail.html",
        context,
    )

# ============================================================
# 奨学生 面談専用画面
# ============================================================

@club_advisor_required
@transaction.atomic
def scholarship_interview(request, student_id):

    teacher = request.teacher

    # --------------------------------------------------------
    # 対象生徒
    #
    # 部顧問
    #   → 自分の担当・登録生徒のみ
    #
    # 広報管理者・システム管理者
    #   → 全対象生徒
    # --------------------------------------------------------

    student_queryset = (
        ProspectiveStudent.objects
        .select_related(
            "junior_high_school",
            "club",
            "assigned_teacher",
            "registered_by",
        )
        .filter(
            is_active=True,
        )
    )

    if teacher.role not in {
        "publicity_admin",
        "system_admin",
    }:

        student_queryset = student_queryset.filter(
            Q(assigned_teacher=teacher)
            | Q(registered_by=teacher)
        )

    student = get_object_or_404(
        student_queryset,
        id=student_id,
    )


    # --------------------------------------------------------
    # 奨学生管理本体
    # --------------------------------------------------------

    assignment, _ = (
        ScholarshipAssignment.objects
        .select_related(
            "interview_rank",
            "current_rank",
            "final_rank",
            "interview_dormitory_benefit",
            "current_dormitory_benefit",
            "final_dormitory_benefit",
        )
        .get_or_create(
            student=student,
        )
    )


    # --------------------------------------------------------
    # 面談前ランク未設定の場合は個人管理画面へ戻す
    # --------------------------------------------------------

    if assignment.current_rank is None:

        messages.warning(
            request,
            "面談を開始する前に、"
            "面談前ランクを設定してください。",
        )

        return redirect(
            "publicity:scholarship_assignment_detail",
            student_id=student.id,
        )


    # --------------------------------------------------------
    # 選択肢
    # --------------------------------------------------------

    scholarship_categories = (
        ScholarshipCategory.objects
        .filter(
            is_active=True,
        )
        .order_by(
            "name",
        )
    )

    dormitory_categories = (
        DormitoryBenefitCategory.objects
        .filter(
            is_active=True,
        )
        .order_by(
            "name",
        )
    )


    # ========================================================
    # POST
    # 面談内容保存
    # ========================================================

    if request.method == "POST":

        # ----------------------------------------------------
        # フォーム値
        # ----------------------------------------------------

        rank_id = (
            request.POST
            .get(
                "presented_rank",
                "",
            )
            .strip()
        )

        dormitory_id = (
            request.POST
            .get(
                "presented_dormitory_benefit",
                "",
            )
            .strip()
        )

        result = (
            request.POST
            .get(
                "result",
                "pending",
            )
            .strip()
        )

        notes = (
            request.POST
            .get(
                "notes",
                "",
            )
            .strip()
        )

        rank_change_reason = (
            request.POST
            .get(
                "rank_change_reason",
                "",
            )
            .strip()
        )

        dormitory_change_reason = (
            request.POST
            .get(
                "dormitory_change_reason",
                "",
            )
            .strip()
        )


        # ----------------------------------------------------
        # 面談結果の値チェック
        # ----------------------------------------------------

        valid_results = {
            value
            for value, label
            in ScholarshipInterview.RESULT_CHOICES
        }

        if result not in valid_results:

            messages.error(
                request,
                "面談結果の値が正しくありません。",
            )

            return redirect(
                "publicity:scholarship_interview",
                student_id=student.id,
            )


        # ----------------------------------------------------
        # 提示ランク取得
        # ----------------------------------------------------

        if not rank_id:

            messages.error(
                request,
                "面談時に提示するランクを"
                "選択してください。",
            )

            return redirect(
                "publicity:scholarship_interview",
                student_id=student.id,
            )


        presented_rank = get_object_or_404(
            ScholarshipCategory,
            id=rank_id,
            is_active=True,
        )


        # ----------------------------------------------------
        # 面談時提示寮費区分
        # 未選択も許可
        # ----------------------------------------------------

        presented_dormitory_benefit = None

        if dormitory_id:

            presented_dormitory_benefit = (
                get_object_or_404(
                    DormitoryBenefitCategory,
                    id=dormitory_id,
                    is_active=True,
                )
            )


        # ====================================================
        # 保存前の現在値を保持
        # ====================================================

        previous_rank = (
            assignment.current_rank
        )

        previous_dormitory_benefit = (
            assignment.current_dormitory_benefit
        )


        # ====================================================
        # ランクが変更された場合
        # ====================================================

        if previous_rank != presented_rank:

            ScholarshipRankHistory.objects.create(
                assignment=assignment,
                previous_rank=previous_rank,
                new_rank=presented_rank,
                reason=(
                    rank_change_reason
                    or "面談時の条件変更"
                ),
                changed_by=teacher,
            )


        # ====================================================
        # 寮費区分が変更された場合
        #
        # new_benefit が null=False なので、
        # 新しい寮費区分が存在するときに履歴を作成。
        # ====================================================

        if (
            previous_dormitory_benefit
            != presented_dormitory_benefit
        ):

            if (
                presented_dormitory_benefit
                is not None
            ):

                DormitoryBenefitHistory.objects.create(
                    assignment=assignment,
                    previous_benefit=(
                        previous_dormitory_benefit
                    ),
                    new_benefit=(
                        presented_dormitory_benefit
                    ),
                    reason=(
                        dormitory_change_reason
                        or "面談時の条件変更"
                    ),
                    changed_by=teacher,
                )


        # ====================================================
        # 面談記録作成
        #
        # ここに保存する presented_rank は
        # 「この面談で実際に提示したランク」
        # ====================================================

        interview = (
            ScholarshipInterview.objects
            .create(
                assignment=assignment,

                interviewed_at=(
                    timezone.now()
                ),

                interviewer=teacher,

                presented_rank=(
                    presented_rank
                ),

                presented_dormitory_benefit=(
                    presented_dormitory_benefit
                ),

                result=result,

                notes=notes,
            )
        )


        # ====================================================
        # ScholarshipAssignment の現在値更新
        #
        # interview_rank は変更しない。
        # 面談前の基準として残す。
        # ====================================================

        assignment.current_rank = (
            presented_rank
        )

        assignment.current_dormitory_benefit = (
            presented_dormitory_benefit
        )


        # ====================================================
        # 面談結果に応じて進行状況更新
        # ====================================================

        if result == "temporary_accepted":

            assignment.status = (
                "temporary_accepted"
            )

            assignment.temporary_accepted = True

            assignment.temporary_accepted_at = (
                timezone.now()
            )


        elif result == "considering":

            assignment.status = (
                "interviewed"
            )

            assignment.temporary_accepted = False

            assignment.temporary_accepted_at = None


        elif result == "declined":

            assignment.status = (
                "declined"
            )

            assignment.temporary_accepted = False

            assignment.temporary_accepted_at = None


        else:

            # pending
            assignment.status = (
                "interviewed"
            )


        assignment.save()


        # ====================================================
        # 完了
        # ====================================================

        messages.success(
            request,
            (
                f"{student.name}さんの"
                "面談内容を保存しました。"
            ),
        )

        return redirect(
            "publicity:scholarship_assignment_detail",
            student_id=student.id,
        )


    # ========================================================
    # GET
    # ========================================================

    context = {

        "teacher": teacher,

        "student": student,

        "assignment": assignment,

        "scholarship_categories": (
            scholarship_categories
        ),

        "dormitory_categories": (
            dormitory_categories
        ),

        "result_choices": (
            ScholarshipInterview.RESULT_CHOICES
        ),
    }


    return render(
        request,
        "publicity/scholarship_interview.html",
        context,
    )

# ============================================================
# 奨学生枠設定
# 管理職・システム管理者向け
# ============================================================

@club_advisor_required
@transaction.atomic
def scholarship_quota_manage(request):

    teacher = request.teacher

    # --------------------------------------------------------
    # 権限
    # --------------------------------------------------------

    if (
        not request.user.is_superuser
        and teacher.role not in {
            "publicity_admin",
            "system_admin",
        }
    ):
        raise PermissionDenied(
            "奨学生枠を設定する権限がありません。"
        )

    # --------------------------------------------------------
    # 年度
    # GET / POST で指定
    # 未指定時は来年度を初期値にする
    # --------------------------------------------------------

    today = timezone.localdate()

    default_fiscal_year = today.year + 1

    fiscal_year_raw = (
        request.POST.get("fiscal_year")
        if request.method == "POST"
        else request.GET.get("fiscal_year")
    )

    try:
        fiscal_year = int(
            fiscal_year_raw
            or default_fiscal_year
        )

    except (TypeError, ValueError):
        fiscal_year = default_fiscal_year

    # --------------------------------------------------------
    # 使用中部活動
    # --------------------------------------------------------

    clubs = (
        Club.objects
        .filter(
            is_active=True
        )
        .order_by(
            "name"
        )
    )

    # --------------------------------------------------------
    # 使用中奨学金区分
    #
    # 表示順：
    # SS → S → A → その他
    # --------------------------------------------------------

    categories = (
        ScholarshipCategory.objects
        .filter(
            is_active=True
        )
        .annotate(
            display_order=Case(
                When(
                    name="SS",
                    then=Value(1),
                ),
                When(
                    name="S",
                    then=Value(2),
                ),
                When(
                    name="A",
                    then=Value(3),
                ),
                default=Value(99),
                output_field=IntegerField(),
            )
        )
        .order_by(
            "display_order",
            "name",
        )
    )

    # --------------------------------------------------------
    # 上限管理対象となる寮費区分
    #
    # DB上の名称と完全一致させる
    # --------------------------------------------------------

    dormitory_quota_benefit = (
        DormitoryBenefitCategory.objects
        .filter(
            is_active=True,
            name="寮施設管理費免除",
        )
        .first()
    )

    # ========================================================
    # POST 保存
    # ========================================================

    if request.method == "POST":

        # ====================================================
        # 1. 奨学生ランク枠を保存
        # ====================================================

        for club in clubs:

            for category in categories:

                field_name = (
                    f"quota_"
                    f"{club.id}_"
                    f"{category.id}"
                )

                raw_value = (
                    request.POST
                    .get(
                        field_name,
                        "",
                    )
                    .strip()
                )

                # 空欄は0として扱う
                if raw_value == "":
                    quota_value = 0

                else:

                    try:
                        quota_value = int(
                            raw_value
                        )

                    except ValueError:

                        messages.error(
                            request,
                            (
                                f"{club.name} "
                                f"{category.name}の"
                                "上限人数が正しくありません。"
                            ),
                        )

                        return redirect(
                            (
                                "publicity:"
                                "scholarship_quota_manage"
                            )
                            + f"?fiscal_year={fiscal_year}"
                        )

                    if quota_value < 0:

                        messages.error(
                            request,
                            (
                                f"{club.name} "
                                f"{category.name}の"
                                "上限人数に負数は指定できません。"
                            ),
                        )

                        return redirect(
                            (
                                "publicity:"
                                "scholarship_quota_manage"
                            )
                            + f"?fiscal_year={fiscal_year}"
                        )

                ScholarshipQuota.objects.update_or_create(
                    fiscal_year=fiscal_year,
                    club=club,
                    category=category,
                    defaults={
                        "quota": quota_value,
                    },
                )

        # ====================================================
        # 2. 寮施設管理費免除枠を保存
        # ====================================================

        if dormitory_quota_benefit is not None:

            for club in clubs:

                field_name = (
                    f"dormitory_quota_{club.id}"
                )

                raw_value = (
                    request.POST
                    .get(
                        field_name,
                        "",
                    )
                    .strip()
                )

                # 空欄は0として扱う
                if raw_value == "":
                    quota_value = 0

                else:

                    try:
                        quota_value = int(
                            raw_value
                        )

                    except ValueError:

                        messages.error(
                            request,
                            (
                                f"{club.name}の"
                                "寮施設管理費免除枠が"
                                "正しくありません。"
                            ),
                        )

                        return redirect(
                            (
                                "publicity:"
                                "scholarship_quota_manage"
                            )
                            + f"?fiscal_year={fiscal_year}"
                        )

                    if quota_value < 0:

                        messages.error(
                            request,
                            (
                                f"{club.name}の"
                                "寮施設管理費免除枠に"
                                "負数は指定できません。"
                            ),
                        )

                        return redirect(
                            (
                                "publicity:"
                                "scholarship_quota_manage"
                            )
                            + f"?fiscal_year={fiscal_year}"
                        )

                DormitoryBenefitQuota.objects.update_or_create(
                    fiscal_year=fiscal_year,
                    club=club,
                    benefit=dormitory_quota_benefit,
                    defaults={
                        "quota": quota_value,
                    },
                )

        # ----------------------------------------------------
        # 寮費区分が見つからなかった場合
        # ----------------------------------------------------

        else:

            messages.warning(
                request,
                (
                    "「寮施設管理費免除」の"
                    "寮費区分が見つからなかったため、"
                    "寮費枠は保存されませんでした。"
                ),
            )

        messages.success(
            request,
            (
                f"{fiscal_year}年度の"
                "奨学生枠を保存しました。"
            ),
        )

        return redirect(
            (
                "publicity:"
                "scholarship_quota_manage"
            )
            + f"?fiscal_year={fiscal_year}"
        )

    # ========================================================
    # GET表示
    # ========================================================

    # --------------------------------------------------------
    # 既存の奨学生ランク枠を取得
    # --------------------------------------------------------

    quotas = (
        ScholarshipQuota.objects
        .filter(
            fiscal_year=fiscal_year,
            club__in=clubs,
            category__in=categories,
        )
        .select_related(
            "club",
            "category",
        )
    )

    # --------------------------------------------------------
    # 奨学生枠を辞書化
    #
    # key:
    # (club_id, category_id)
    #
    # value:
    # quota
    # --------------------------------------------------------

    quota_map = {
        (
            quota.club_id,
            quota.category_id,
        ): quota.quota

        for quota in quotas
    }

    # --------------------------------------------------------
    # 既存の寮費枠を取得
    # --------------------------------------------------------

    dormitory_quota_map = {}

    if dormitory_quota_benefit is not None:

        dormitory_quotas = (
            DormitoryBenefitQuota.objects
            .filter(
                fiscal_year=fiscal_year,
                club__in=clubs,
                benefit=dormitory_quota_benefit,
            )
            .select_related(
                "club",
                "benefit",
            )
        )

        # ----------------------------------------------------
        # key:
        # club_id
        #
        # value:
        # quota
        # ----------------------------------------------------

        dormitory_quota_map = {
            quota.club_id: quota.quota
            for quota in dormitory_quotas
        }

    # --------------------------------------------------------
    # Template用行データ
    #
    # Django templateでは
    # tuple keyの辞書を扱いにくいため
    # View側で表示用データを完成させる
    # --------------------------------------------------------

    quota_rows = []

    for club in clubs:

        values = []

        # ----------------------------------------------------
        # SS / S / A
        # ----------------------------------------------------

        for category in categories:

            values.append(
                {
                    "category": category,

                    "value": quota_map.get(
                        (
                            club.id,
                            category.id,
                        ),
                        0,
                    ),

                    "field_name": (
                        f"quota_"
                        f"{club.id}_"
                        f"{category.id}"
                    ),
                }
            )

        # ----------------------------------------------------
        # 1部活動分を作成
        # ----------------------------------------------------

        quota_rows.append(
            {
                "club": club,

                # SS / S / A
                "values": values,

                # 寮施設管理費免除
                "dormitory_value":
                    dormitory_quota_map.get(
                        club.id,
                        0,
                    ),

                "dormitory_field_name":
                    f"dormitory_quota_{club.id}",
            }
        )

    # --------------------------------------------------------
    # Templateへ渡す
    # --------------------------------------------------------

    context = {
        "teacher": teacher,

        "fiscal_year":
            fiscal_year,

        "categories":
            categories,

        "quota_rows":
            quota_rows,

        "dormitory_quota_benefit":
            dormitory_quota_benefit,
    }

    return render(
        request,
        "publicity/scholarship_quota_manage.html",
        context,
    )

# ============================================================
# 募集連絡・面談進捗管理
# ============================================================

@club_advisor_required
def recruitment_response_management_list(request):
    """
    募集対象生徒について、

    ・中学校への連絡状況
    ・部顧問の返答
    ・面談状況
    ・面談結果

    を一覧管理する画面。

    広報管理者・システム管理者：
        全対象生徒を表示

    部顧問：
        自分が登録または担当している生徒のみ表示
    """

    teacher = request.teacher

    # ============================================================
    # 基本QuerySet
    # ============================================================

    base_students = (
        ProspectiveStudent.objects
        .filter(
            is_active=True
        )
        .select_related(
            "junior_high_school",
            "club",
            "assigned_teacher",
            "registered_by",
            "scholarship_category",
        )
    )

    # ============================================================
    # 権限制御
    # ============================================================

    if teacher.role not in {
        "publicity_admin",
        "system_admin",
    }:

        base_students = (
            base_students.filter(
                Q(
                    registered_by=teacher
                )
                | Q(
                    assigned_teacher=teacher
                )
            )
            .distinct()
        )

    # ============================================================
    # 集計
    #
    # 検索前の全体進捗を表示する
    # ============================================================

    total_count = (
        base_students.count()
    )

    not_contacted_count = (
        base_students.filter(
            contact_status="not_contacted"
        )
        .count()
    )

    waiting_count = (
        base_students.filter(
            contact_status="waiting"
        )
        .count()
    )

    responded_count = (
        base_students.filter(
            contact_status="responded"
        )
        .count()
    )

    advisor_accept_count = (
        base_students.filter(
            advisor_response="accept"
        )
        .count()
    )

    advisor_decline_count = (
        base_students.filter(
            advisor_response="decline"
        )
        .count()
    )

    interview_scheduled_count = (
        base_students.filter(
            interview_status="scheduled"
        )
        .count()
    )

    interview_completed_count = (
        base_students.filter(
            interview_status="completed"
        )
        .count()
    )

    candidate_count = (
        base_students.filter(
            interview_result="candidate"
        )
        .count()
    )

    # ============================================================
    # GET検索条件
    # ============================================================

    keyword = request.GET.get(
        "q",
        "",
    ).strip()

    school_id = request.GET.get(
        "school",
        "",
    ).strip()

    club_id = request.GET.get(
        "club",
        "",
    ).strip()

    contact_status = request.GET.get(
        "contact_status",
        "",
    ).strip()

    advisor_response = request.GET.get(
        "advisor_response",
        "",
    ).strip()

    interview_status = request.GET.get(
        "interview_status",
        "",
    ).strip()

    interview_result = request.GET.get(
        "interview_result",
        "",
    ).strip()

    # ============================================================
    # 検索対象QuerySet
    # ============================================================

    students = base_students

    # ------------------------------------------------------------
    # キーワード
    # ------------------------------------------------------------

    if keyword:

        students = students.filter(
            Q(
                name__icontains=keyword
            )
            | Q(
                name_kana__icontains=keyword
            )
            | Q(
                junior_high_school__name__icontains=keyword
            )
            | Q(
                club__name__icontains=keyword
            )
            | Q(
                assigned_teacher__name__icontains=keyword
            )
        )

    # ------------------------------------------------------------
    # 中学校
    # ------------------------------------------------------------

    if school_id.isdigit():

        students = students.filter(
            junior_high_school_id=school_id
        )

    # ------------------------------------------------------------
    # 部活動
    # ------------------------------------------------------------

    if club_id.isdigit():

        students = students.filter(
            club_id=club_id
        )

    # ------------------------------------------------------------
    # 中学校連絡状況
    # ------------------------------------------------------------

    if contact_status:

        students = students.filter(
            contact_status=contact_status
        )

    # ------------------------------------------------------------
    # 部顧問返答
    # ------------------------------------------------------------

    if advisor_response:

        students = students.filter(
            advisor_response=advisor_response
        )

    # ------------------------------------------------------------
    # 面談状況
    # ------------------------------------------------------------

    if interview_status:

        students = students.filter(
            interview_status=interview_status
        )

    # ------------------------------------------------------------
    # 面談結果
    # ------------------------------------------------------------

    if interview_result:

        students = students.filter(
            interview_result=interview_result
        )

    # ============================================================
    # 地域優先順
    #
    # 小林 → えびの → 高原 → 都城 → 三股 → 宮崎
    # ============================================================

    students = (
        students.annotate(
            area_order=Case(

                When(
                    junior_high_school__city__icontains="小林",
                    then=Value(1),
                ),

                When(
                    junior_high_school__city__icontains="えびの",
                    then=Value(2),
                ),

                When(
                    junior_high_school__city__icontains="高原",
                    then=Value(3),
                ),

                When(
                    junior_high_school__city__icontains="都城",
                    then=Value(4),
                ),

                When(
                    junior_high_school__city__icontains="三股",
                    then=Value(5),
                ),

                When(
                    junior_high_school__city__icontains="宮崎",
                    then=Value(6),
                ),

                When(
                    junior_high_school__prefecture__icontains="宮崎",
                    then=Value(90),
                ),

                default=Value(99),

                output_field=IntegerField(),
            )
        )
        .order_by(
            "area_order",
            "junior_high_school__city",
            "junior_high_school__name",
            "club__name",
            "name_kana",
            "name",
        )
    )

    # ============================================================
    # 中学校選択肢
    # ============================================================

    school_ids = (
        base_students
        .values_list(
            "junior_high_school_id",
            flat=True,
        )
        .distinct()
    )

    schools = (
        JuniorHighSchool.objects
        .filter(
            id__in=school_ids
        )
        .order_by(
            "city",
            "name",
        )
    )

    # ============================================================
    # 部活動選択肢
    # ============================================================

    club_ids = (
        base_students
        .values_list(
            "club_id",
            flat=True,
        )
        .distinct()
    )

    clubs = (
        Club.objects
        .filter(
            id__in=club_ids
        )
        .order_by(
            "name"
        )
    )

    # ============================================================
    # Context
    # ============================================================

    context = {

        "students": students,

        "schools": schools,

        "clubs": clubs,

        # --------------------------------------------------------
        # 検索条件
        # --------------------------------------------------------

        "keyword": keyword,

        "selected_school": (
            school_id
        ),

        "selected_club": (
            club_id
        ),

        "selected_contact_status": (
            contact_status
        ),

        "selected_advisor_response": (
            advisor_response
        ),

        "selected_interview_status": (
            interview_status
        ),

        "selected_interview_result": (
            interview_result
        ),

        # --------------------------------------------------------
        # choices
        # --------------------------------------------------------

        "contact_status_choices": (
            ProspectiveStudent
            .CONTACT_STATUS_CHOICES
        ),

        "advisor_response_choices": (
            ProspectiveStudent
            .ADVISOR_RESPONSE_CHOICES
        ),

        "interview_status_choices": (
            ProspectiveStudent
            .INTERVIEW_STATUS_CHOICES
        ),

        "interview_result_choices": (
            ProspectiveStudent
            .INTERVIEW_RESULT_CHOICES
        ),

        # --------------------------------------------------------
        # 集計
        # --------------------------------------------------------

        "total_count": (
            total_count
        ),

        "not_contacted_count": (
            not_contacted_count
        ),

        "waiting_count": (
            waiting_count
        ),

        "responded_count": (
            responded_count
        ),

        "advisor_accept_count": (
            advisor_accept_count
        ),

        "advisor_decline_count": (
            advisor_decline_count
        ),

        "interview_scheduled_count": (
            interview_scheduled_count
        ),

        "interview_completed_count": (
            interview_completed_count
        ),

        "candidate_count": (
            candidate_count
        ),

        "display_count": (
            students.count()
        ),
    }

    return render(
        request,
        (
            "publicity/"
            "recruitment_response_management_list.html"
        ),
        context,
    )

# ============================================================
# 募集連絡・面談 個人管理
# ============================================================

@club_advisor_required
@transaction.atomic
def recruitment_response_management_detail(
    request,
    student_id,
):

    teacher = request.teacher

    student = get_object_or_404(
        ProspectiveStudent.objects
        .select_related(
            "junior_high_school",
            "club",
            "assigned_teacher",
            "registered_by",
            "scholarship_category",
        ),
        pk=student_id,
        is_active=True,
    )

    # ============================================================
    # 権限制御
    # ============================================================

    if teacher.role not in {
        "publicity_admin",
        "system_admin",
    }:

        if (
            student.registered_by_id != teacher.id
            and student.assigned_teacher_id != teacher.id
        ):
            raise PermissionDenied(
                "この生徒の連絡・面談情報を管理する権限がありません。"
            )

    # ============================================================
    # POST：更新
    # ============================================================

    if request.method == "POST":

        contact_status = request.POST.get(
            "contact_status",
            "",
        ).strip()

        contact_date_text = request.POST.get(
            "contact_date",
            "",
        ).strip()

        contact_memo = request.POST.get(
            "contact_memo",
            "",
        ).strip()

        advisor_response = request.POST.get(
            "advisor_response",
            "",
        ).strip()

        advisor_response_memo = request.POST.get(
            "advisor_response_memo",
            "",
        ).strip()

        interview_status = request.POST.get(
            "interview_status",
            "",
        ).strip()

        interview_date_text = request.POST.get(
            "interview_date",
            "",
        ).strip()

        interview_memo = request.POST.get(
            "interview_memo",
            "",
        ).strip()

        interview_result = request.POST.get(
            "interview_result",
            "",
        ).strip()

        management_memo = request.POST.get(
            "management_memo",
            "",
        ).strip()

        scholarship_category_id = request.POST.get(
            "scholarship_category",
            "",
        ).strip()

        # ========================================================
        # choices 検証
        # ========================================================

        valid_contact_statuses = {
            value
            for value, label
            in ProspectiveStudent.CONTACT_STATUS_CHOICES
        }

        valid_advisor_responses = {
            value
            for value, label
            in ProspectiveStudent.ADVISOR_RESPONSE_CHOICES
        }

        valid_interview_statuses = {
            value
            for value, label
            in ProspectiveStudent.INTERVIEW_STATUS_CHOICES
        }

        valid_interview_results = {
            value
            for value, label
            in ProspectiveStudent.INTERVIEW_RESULT_CHOICES
        }

        if contact_status not in valid_contact_statuses:

            messages.error(
                request,
                "中学校連絡状況が正しくありません。",
            )

            return redirect(
                "publicity:recruitment_response_management_detail",
                student_id=student.id,
            )

        if advisor_response not in valid_advisor_responses:

            messages.error(
                request,
                "部顧問返答が正しくありません。",
            )

            return redirect(
                "publicity:recruitment_response_management_detail",
                student_id=student.id,
            )

        if interview_status not in valid_interview_statuses:

            messages.error(
                request,
                "面談状況が正しくありません。",
            )

            return redirect(
                "publicity:recruitment_response_management_detail",
                student_id=student.id,
            )

        if interview_result not in valid_interview_results:

            messages.error(
                request,
                "面談結果が正しくありません。",
            )

            return redirect(
                "publicity:recruitment_response_management_detail",
                student_id=student.id,
            )

        # ========================================================
        # 日付
        # ========================================================

        contact_date = None

        if contact_date_text:

            try:
                contact_date = datetime.strptime(
                    contact_date_text,
                    "%Y-%m-%d",
                ).date()

            except ValueError:

                messages.error(
                    request,
                    "中学校連絡日の形式が正しくありません。",
                )

                return redirect(
                    "publicity:recruitment_response_management_detail",
                    student_id=student.id,
                )

        interview_date = None

        if interview_date_text:

            try:
                interview_date = datetime.strptime(
                    interview_date_text,
                    "%Y-%m-%d",
                ).date()

            except ValueError:

                messages.error(
                    request,
                    "面談日の形式が正しくありません。",
                )

                return redirect(
                    "publicity:recruitment_response_management_detail",
                    student_id=student.id,
                )

        # ========================================================
        # 奨学金区分
        # ========================================================

        scholarship_category = None

        if scholarship_category_id.isdigit():

            scholarship_category = get_object_or_404(
                ScholarshipCategory,
                pk=scholarship_category_id,
                is_active=True,
            )

        # ========================================================
        # 保存
        # ========================================================

        student.contact_status = (
            contact_status
        )

        student.contact_date = (
            contact_date
        )

        student.contact_memo = (
            contact_memo
        )

        student.advisor_response = (
            advisor_response
        )

        student.advisor_response_memo = (
            advisor_response_memo
        )

        student.interview_status = (
            interview_status
        )

        student.interview_date = (
            interview_date
        )

        student.interview_memo = (
            interview_memo
        )

        student.interview_result = (
            interview_result
        )

        student.management_memo = (
            management_memo
        )

        student.scholarship_category = (
            scholarship_category
        )

        student.save(
            update_fields=[
                "contact_status",
                "contact_date",
                "contact_memo",
                "advisor_response",
                "advisor_response_memo",
                "interview_status",
                "interview_date",
                "interview_memo",
                "interview_result",
                "management_memo",
                "scholarship_category",
                "updated_at",
            ]
        )

        messages.success(
            request,
            f"{student.name}さんの連絡・面談情報を更新しました。",
        )

        return redirect(
            "publicity:recruitment_response_management_detail",
            student_id=student.id,
        )

    # ============================================================
    # 奨学金区分
    # ============================================================

    scholarship_categories = (
        ScholarshipCategory.objects
        .filter(
            is_active=True
        )
        .annotate(
            display_order=Case(
                When(
                    name="SS",
                    then=Value(1),
                ),
                When(
                    name="S",
                    then=Value(2),
                ),
                When(
                    name="A",
                    then=Value(3),
                ),
                default=Value(99),
                output_field=IntegerField(),
            )
        )
        .order_by(
            "display_order",
            "name",
        )
    )

    # ============================================================
    # Context
    # ============================================================

    context = {

        "student": student,

        "contact_status_choices": (
            ProspectiveStudent
            .CONTACT_STATUS_CHOICES
        ),

        "advisor_response_choices": (
            ProspectiveStudent
            .ADVISOR_RESPONSE_CHOICES
        ),

        "interview_status_choices": (
            ProspectiveStudent
            .INTERVIEW_STATUS_CHOICES
        ),

        "interview_result_choices": (
            ProspectiveStudent
            .INTERVIEW_RESULT_CHOICES
        ),

        "scholarship_categories": (
            scholarship_categories
        ),
    }

    return render(
        request,
        (
            "publicity/"
            "recruitment_response_management_detail.html"
        ),
        context,
    )

@club_advisor_required
@transaction.atomic
def junior_high_school_quick_create(request):
    """
    募集対象生徒登録画面から、
    文科省マスタに存在しない中学校を手動登録する。
    """

    if request.method != "POST":
        return JsonResponse(
            {
                "success": False,
                "message": "POSTで送信してください。",
            },
            status=405,
        )

    # ============================================================
    # 入力値
    # ============================================================

    name = request.POST.get(
        "name",
        "",
    ).strip()

    prefecture = request.POST.get(
        "prefecture",
        "",
    ).strip()

    city = request.POST.get(
        "city",
        "",
    ).strip()

    principal_name = request.POST.get(
        "principal_name",
        "",
    ).strip()

    tel = request.POST.get(
        "tel",
        "",
    ).strip()

    address = request.POST.get(
        "address",
        "",
    ).strip()

    # ============================================================
    # 必須チェック
    # ============================================================

    if not name:
        return JsonResponse(
            {
                "success": False,
                "message": "学校名を入力してください。",
            },
            status=400,
        )

    if not prefecture:
        return JsonResponse(
            {
                "success": False,
                "message": "都道府県を入力してください。",
            },
            status=400,
        )

    if not city:
        return JsonResponse(
            {
                "success": False,
                "message": "市町村を入力してください。",
            },
            status=400,
        )

    # ============================================================
    # 重複チェック
    #
    # 同じ都道府県・市町村・学校名がすでにある場合は
    # 新規作成せず、その学校を返す。
    # ============================================================

    existing_school = (
        JuniorHighSchool.objects
        .filter(
            name__iexact=name,
            prefecture__iexact=prefecture,
            city__iexact=city,
        )
        .first()
    )

    if existing_school:

        # 無効になっていた場合は復活
        if not existing_school.is_active:
            existing_school.is_active = True
            existing_school.save(
                update_fields=[
                    "is_active",
                    "updated_at",
                ]
            )

        return JsonResponse(
            {
                "success": True,
                "created": False,
                "school": {
                    "id": existing_school.id,
                    "name": existing_school.name,
                    "prefecture": (
                        existing_school.prefecture
                        or ""
                    ),
                    "city": (
                        existing_school.city
                        or ""
                    ),
                    "principal_name": (
                        existing_school.principal_name
                        or ""
                    ),
                },
                "message": (
                    "すでに登録されている中学校を"
                    "選択しました。"
                ),
            },
            json_dumps_params={
                "ensure_ascii": False,
            },
        )

    # ============================================================
    # 新規登録
    # ============================================================

    school = JuniorHighSchool.objects.create(
        name=name,
        prefecture=prefecture,
        city=city,
        principal_name=principal_name,
        tel=tel,
        address=address,

        # 文科省データではなく手動登録
        is_from_mext=False,

        is_active=True,
    )

    return JsonResponse(
        {
            "success": True,
            "created": True,
            "school": {
                "id": school.id,
                "name": school.name,
                "prefecture": (
                    school.prefecture
                    or ""
                ),
                "city": (
                    school.city
                    or ""
                ),
                "principal_name": (
                    school.principal_name
                    or ""
                ),
            },
            "message": (
                f"{school.name}を登録しました。"
            ),
        },
        json_dumps_params={
            "ensure_ascii": False,
        },
    )
# ============================================================
# 宮崎県内 中学校資料 持参用人数リストPDF
# ============================================================

@publicity_admin_required
def miyazaki_material_delivery_list_pdf(request):
    """
    宮崎県内の選択した中学校について、
    資料持参・封入作業用の一覧PDFを作成する。

    クラス数 ＝ 必要な封筒数
    各クラスの total ＝ 各封筒へ入れる資料数

    並び順：
    小林市
    → 高原町
    → えびの市
    → 都城市
    → 三股町
    → 宮崎市
    → その他宮崎県内
    """

    # ========================================================
    # 選択学校ID
    # ========================================================

    school_ids = request.GET.getlist(
        "school_ids"
    )

    if not school_ids:
        return HttpResponse(
            "中学校を1校以上選択してください。",
            status=400,
        )

    # ========================================================
    # 対象中学校
    # ========================================================

    schools = (
        JuniorHighSchool.objects
        .filter(
            id__in=school_ids,
            is_active=True,
            prefecture="宮崎県",
        )
        .prefetch_related(
            "classes"
        )
        .annotate(
            area_order=Case(

                When(
                    city__icontains="小林",
                    then=Value(1),
                ),

                When(
                    city__icontains="高原",
                    then=Value(2),
                ),

                When(
                    city__icontains="えびの",
                    then=Value(3),
                ),

                When(
                    city__icontains="都城",
                    then=Value(4),
                ),

                When(
                    city__icontains="三股",
                    then=Value(5),
                ),

                When(
                    city__icontains="宮崎",
                    then=Value(6),
                ),

                default=Value(99),

                output_field=IntegerField(),
            )
        )
        .order_by(
            "area_order",
            "city",
            "name",
        )
    )

    schools = list(schools)

    if not schools:
        return HttpResponse(
            "対象となる宮崎県内の中学校がありません。",
            status=400,
        )

    # ========================================================
    # 市町村単位でグループ化
    # ========================================================

    grouped_schools = {}

    for school in schools:

        city_name = (
            school.city.strip()
            if school.city
            else "市町村未設定"
        )

        if city_name not in grouped_schools:
            grouped_schools[city_name] = []

        grouped_schools[
            city_name
        ].append(
            school
        )

    # ========================================================
    # PDF準備
    # ========================================================

    buffer = BytesIO()

    pdf = canvas.Canvas(
        buffer,
        pagesize=A4,
    )

    width, height = A4

    # ========================================================
    # 日本語フォント
    # ========================================================

    font_min = "HeiseiMin-W3"
    font_gothic = "HeiseiKakuGo-W5"

    pdfmetrics.registerFont(
        UnicodeCIDFont(
            font_min
        )
    )

    pdfmetrics.registerFont(
        UnicodeCIDFont(
            font_gothic
        )
    )

    # ========================================================
    # レイアウト
    # ========================================================

    left_margin = 18 * mm
    right_margin = 18 * mm

    top_margin = 18 * mm
    bottom_margin = 16 * mm

    content_width = (
        width
        - left_margin
        - right_margin
    )

    page_number = 1

    # ========================================================
    # ページヘッダー
    # ========================================================

    def draw_page_header():

        pdf.setFont(
            font_gothic,
            16,
        )

        pdf.drawCentredString(
            width / 2,
            height - top_margin,
            "宮崎県内 中学校資料持参・封入一覧",
        )

        pdf.setFont(
            font_min,
            9.5,
        )

        pdf.drawString(
            left_margin,
            height - top_margin - 9 * mm,
            (
                "クラスごとに封筒を用意し、"
                "記載人数分の資料を封入してください。"
            ),
        )

        today = timezone.localdate()

        reiwa_year = (
            today.year - 2018
        )

        date_label = (
            f"令和{reiwa_year}年"
            f"{today.month}月"
            f"{today.day}日"
        )

        pdf.drawRightString(
            width - right_margin,
            height - top_margin - 9 * mm,
            date_label,
        )

        pdf.setFont(
            font_min,
            8.5,
        )

        pdf.drawRightString(
            width - right_margin,
            bottom_margin - 5 * mm,
            f"{page_number}ページ",
        )

        return (
            height
            - top_margin
            - 18 * mm
        )

    # ========================================================
    # 改ページ
    # ========================================================

    def new_page():

        nonlocal page_number

        pdf.showPage()

        page_number += 1

        return draw_page_header()

    # ========================================================
    # 最初のページ
    # ========================================================

    current_y = (
        draw_page_header()
    )

    # ========================================================
    # サイズ設定
    # ========================================================

    city_header_height = 10 * mm
    school_header_height = 10 * mm
    table_header_height = 8 * mm
    class_row_height = 8 * mm
    summary_height = 9 * mm

    # ========================================================
    # 市町村単位
    # ========================================================

    for city_name, city_schools in (
        grouped_schools.items()
    ):

        # 市町村見出し＋最低限の学校情報が
        # 入らない場合は改ページ
        if (
            current_y - 42 * mm
            < bottom_margin
        ):
            current_y = new_page()

        # ====================================================
        # 市町村見出し
        # ====================================================

        pdf.setFillColor(
            colors.HexColor(
                "#DDEBF2"
            )
        )

        pdf.roundRect(
            left_margin,
            current_y - city_header_height,
            content_width,
            city_header_height,
            2 * mm,
            fill=1,
            stroke=0,
        )

        pdf.setFillColor(
            colors.black
        )

        pdf.setFont(
            font_gothic,
            12,
        )

        pdf.drawString(
            left_margin + 4 * mm,
            current_y - 6.8 * mm,
            (
                f"【{city_name}】"
                f"　{len(city_schools)}校"
            ),
        )

        current_y -= (
            city_header_height
            + 4 * mm
        )

        # ====================================================
        # 学校単位
        # ====================================================

        for school in city_schools:

            classes = list(
                school.classes.all()
            )

            class_count = len(
                classes
            )

            # ------------------------------------------------
            # クラス人数合計
            # ------------------------------------------------

            class_total = sum(
                cls.total or 0
                for cls in classes
            )

            # ------------------------------------------------
            # 必要高さ
            # ------------------------------------------------

            if classes:

                school_required_height = (
                    school_header_height
                    + table_header_height
                    + (
                        class_row_height
                        * class_count
                    )
                    + summary_height
                    + 6 * mm
                )

            else:

                school_required_height = (
                    school_header_height
                    + 20 * mm
                )

            # ------------------------------------------------
            # 1校丸ごと次ページに送れる場合
            # ------------------------------------------------

            if (
                current_y
                - school_required_height
                < bottom_margin
            ):

                current_y = new_page()

            # =================================================
            # 学校名
            # =================================================

            pdf.setFillColor(
                colors.HexColor(
                    "#F3F3F3"
                )
            )

            pdf.rect(
                left_margin,
                current_y - school_header_height,
                content_width,
                school_header_height,
                fill=1,
                stroke=0,
            )

            pdf.setFillColor(
                colors.black
            )

            # チェックボックス
            checkbox_size = (
                4.5 * mm
            )

            checkbox_x = (
                left_margin
                + 4 * mm
            )

            checkbox_y = (
                current_y
                - school_header_height
                + (
                    school_header_height
                    - checkbox_size
                ) / 2
            )

            pdf.rect(
                checkbox_x,
                checkbox_y,
                checkbox_size,
                checkbox_size,
                fill=0,
                stroke=1,
            )

            pdf.setFont(
                font_gothic,
                11,
            )

            pdf.drawString(
                left_margin + 13 * mm,
                current_y - 6.8 * mm,
                school.name,
            )

            # ------------------------------------------------
            # 学校右側に封筒数
            # ------------------------------------------------

            pdf.setFont(
                font_gothic,
                10.5,
            )

            if classes:

                envelope_label = (
                    f"封筒 {class_count}枚"
                )

            else:

                envelope_label = (
                    "クラス情報未登録"
                )

            pdf.drawRightString(
                width - right_margin - 4 * mm,
                current_y - 6.8 * mm,
                envelope_label,
            )

            current_y -= (
                school_header_height
            )

            # =================================================
            # クラス情報なし
            # =================================================

            if not classes:

                pdf.setFont(
                    font_min,
                    10,
                )

                pdf.setFillColor(
                    colors.HexColor(
                        "#AA0000"
                    )
                )

                pdf.drawString(
                    left_margin + 8 * mm,
                    current_y - 7 * mm,
                    (
                        "※ クラス別人数が"
                        "登録されていません。"
                    ),
                )

                pdf.setFillColor(
                    colors.black
                )

                pdf.drawRightString(
                    width - right_margin - 8 * mm,
                    current_y - 7 * mm,
                    (
                        "3年生合計："
                        f"{school.third_grade_total or 0}名"
                    ),
                )

                current_y -= (
                    14 * mm
                )

                continue

            # =================================================
            # クラス表
            # =================================================

            # 列幅
            class_name_width = (
                70 * mm
            )

            student_width = (
                42 * mm
            )

            instruction_width = (
                content_width
                - class_name_width
                - student_width
            )

            x0 = left_margin
            x1 = (
                x0
                + class_name_width
            )
            x2 = (
                x1
                + student_width
            )

            # =================================================
            # 表見出し
            # =================================================

            pdf.setFillColor(
                colors.HexColor(
                    "#FAFAFA"
                )
            )

            pdf.rect(
                x0,
                current_y - table_header_height,
                content_width,
                table_header_height,
                fill=1,
                stroke=1,
            )

            pdf.setFillColor(
                colors.black
            )

            pdf.line(
                x1,
                current_y,
                x1,
                current_y - table_header_height,
            )

            pdf.line(
                x2,
                current_y,
                x2,
                current_y - table_header_height,
            )

            pdf.setFont(
                font_gothic,
                9.5,
            )

            text_y = (
                current_y - 5.7 * mm
            )

            pdf.drawCentredString(
                x0 + class_name_width / 2,
                text_y,
                "クラス",
            )

            pdf.drawCentredString(
                x1 + student_width / 2,
                text_y,
                "人数",
            )

            pdf.drawCentredString(
                x2 + instruction_width / 2,
                text_y,
                "封入数",
            )

            current_y -= (
                table_header_height
            )

            # =================================================
            # 各クラス
            # =================================================

            for cls in classes:

                # ページ下端対策
                if (
                    current_y
                    - class_row_height
                    < bottom_margin
                ):

                    current_y = new_page()

                    pdf.setFont(
                        font_gothic,
                        10.5,
                    )

                    pdf.drawString(
                        left_margin,
                        current_y,
                        (
                            f"{school.name}"
                            "（続き）"
                        ),
                    )

                    current_y -= (
                        7 * mm
                    )

                row_bottom = (
                    current_y
                    - class_row_height
                )

                # 外枠
                pdf.rect(
                    x0,
                    row_bottom,
                    content_width,
                    class_row_height,
                    fill=0,
                    stroke=1,
                )

                pdf.line(
                    x1,
                    current_y,
                    x1,
                    row_bottom,
                )

                pdf.line(
                    x2,
                    current_y,
                    x2,
                    row_bottom,
                )

                # ------------------------------------------------
                # クラス名
                # ------------------------------------------------

                pdf.setFont(
                    font_min,
                    10.5,
                )

                pdf.drawCentredString(
                    x0 + class_name_width / 2,
                    row_bottom + 2.8 * mm,
                    cls.class_name,
                )

                # ------------------------------------------------
                # 人数
                # ------------------------------------------------

                total = (
                    cls.total or 0
                )

                pdf.drawCentredString(
                    x1 + student_width / 2,
                    row_bottom + 2.8 * mm,
                    f"{total}名",
                )

                # ------------------------------------------------
                # 封入数
                # ------------------------------------------------

                pdf.setFont(
                    font_gothic,
                    10.5,
                )

                pdf.drawCentredString(
                    x2 + instruction_width / 2,
                    row_bottom + 2.8 * mm,
                    f"{total}部",
                )

                current_y -= (
                    class_row_height
                )

            # =================================================
            # 学校合計
            # =================================================

            summary_bottom = (
                current_y
                - summary_height
            )

            pdf.setFillColor(
                colors.HexColor(
                    "#F7F7F7"
                )
            )

            pdf.rect(
                left_margin,
                summary_bottom,
                content_width,
                summary_height,
                fill=1,
                stroke=1,
            )

            pdf.setFillColor(
                colors.black
            )

            pdf.setFont(
                font_gothic,
                10.5,
            )

            pdf.drawString(
                left_margin + 8 * mm,
                summary_bottom + 3 * mm,
                (
                    f"封筒：{class_count}枚"
                ),
            )

            pdf.drawRightString(
                width - right_margin - 8 * mm,
                summary_bottom + 3 * mm,
                (
                    f"資料合計：{class_total}部"
                ),
            )

            current_y -= (
                summary_height
                + 6 * mm
            )

        # ====================================================
        # 市町村間スペース
        # ====================================================

        current_y -= (
            4 * mm
        )

    # ========================================================
    # 最終集計
    # ========================================================

    total_envelopes = 0
    total_materials = 0

    for school in schools:

        classes = list(
            school.classes.all()
        )

        total_envelopes += len(
            classes
        )

        total_materials += sum(
            cls.total or 0
            for cls in classes
        )

    # ========================================================
    # 全体集計を表示
    # ========================================================

    if (
        current_y - 24 * mm
        < bottom_margin
    ):

        current_y = new_page()

    pdf.setFillColor(
        colors.HexColor(
            "#E8F3E8"
        )
    )

    pdf.roundRect(
        left_margin,
        current_y - 17 * mm,
        content_width,
        17 * mm,
        2 * mm,
        fill=1,
        stroke=1,
    )

    pdf.setFillColor(
        colors.black
    )

    pdf.setFont(
        font_gothic,
        12,
    )

    pdf.drawString(
        left_margin + 7 * mm,
        current_y - 7 * mm,
        "全体集計",
    )

    pdf.setFont(
        font_gothic,
        11,
    )

    pdf.drawString(
        left_margin + 45 * mm,
        current_y - 7 * mm,
        (
            f"学校：{len(schools)}校"
        ),
    )

    pdf.drawString(
        left_margin + 85 * mm,
        current_y - 7 * mm,
        (
            f"封筒：{total_envelopes}枚"
        ),
    )

    pdf.drawString(
        left_margin + 125 * mm,
        current_y - 7 * mm,
        (
            f"資料：{total_materials}部"
        ),
    )

    # ========================================================
    # PDF保存
    # ========================================================

    pdf.save()

    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/pdf",
    )

    filename = (
        "宮崎県内_中学校資料持参封入一覧.pdf"
    )

    response[
        "Content-Disposition"
    ] = (
        f'inline; filename="{quote(filename)}"'
    )

    return response